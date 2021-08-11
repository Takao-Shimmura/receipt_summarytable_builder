#!python3.9.1
from flask import Flask, render_template, request, session, \
    redirect, jsonify, current_app, g,send_file
#import psycopg2
import openpyxl # 外部ライブラリ　pip install openpyxl
import pandas as pd
#import numpy as np
import sqlite3
import json
from datetime import datetime

import os
import pathlib
import pytz
import pprint

from sqlalchemy import create_engine, Column, Integer, String, \
    Text, DateTime, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy.sql.elements import Null

from myutil import User,Calculate,ErrorMsg,Search_condition,InsurerData,\
    get_dic_schCond2calAttr,get_search_condition,get_by_list,\
        get_cellno_2list,define_soukatsu1Desti,sort_insureName_4Sokatsu1,\
        soukatsu1Desti_List_set,get_soukatsu1Desti_count,get_soukatsu1Desti_insur_dic,\
        kensuu_insert,kingaku_insert,koukikourei_No_Sort,\
        error_Msg_Sheet


#　↓　herokuにデプロイすると、画像が読み込めない。
# これを解消するためにflaskがheroku内でインスタンス化される時に、
# 静的なファイルのディレクトリを記述して明確化する。
# 参考⇒https://qiita.com/go_new_innov/items/222a3ed92f5ed093f462
app = Flask(__name__,static_folder='./static')

app.secret_key = b'random string...'

#　↓　herokuのpostgreSQL接続用URI 
# ※ただし、割り当てられたURIそのままでは接続エラー
#　「postgres://・・・」から「postgresql://・・・」に変更しなければ解消されない
#参考（heroku公式リファレンス）⇒Why is SQLAlchemy 1.4.x not connecting to Heroku Postgres? - Heroku Help
engine = create_engine('postgresql://qrnkdpytaiifps:7b728dc1e568e2d1c1ab80c919e17d10c7f41f8d853c8e5989d907c978bf8d8c@ec2-34-250-16-127.eu-west-1.compute.amazonaws.com:5432/d77prcb2vt5pne')

#　↓　ローカルのSQLite接続用パス 
#engine = create_engine('sqlite:///sample.sqlite3')

df_new={} #グローバル変数として追加　
#pandasを用いてエクセルを読み込んで作成されたデータフレーム
#を、辞書として整理したものを入れておく変数

# access top page.
@app.route('/',methods=['GET'])
def index():
     #'soukatsuTemp.xlsx'以外のエクセルファイルがもしサーバー上に
     #残っていたときに、HPを更新した時に予め削除してリセットしておく
    path = pathlib.Path("./")   
    for pass_obj in path.iterdir():
        if pass_obj.match("*.xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
            pass_obj.unlink()
    global df_new
    #pprint.pprint('pre df_new={}'.format(df_new))
    
    df_new={}

    #データベースもリセット
    Session = sessionmaker(bind=engine)
    ses = Session()

    ses.query(Calculate).delete()# 復帰
    ses.commit()# 復帰
    ses.close()# 復帰

    ses.query(ErrorMsg).delete() #復帰
    ses.commit()#復帰
    ses.close()# 復帰
    return render_template('index.html',\
            title = '新潟県鍼灸マッサージ師会　公認',\
            message = '保険申請書　総括票作成　ホームページ')


#get calculate
@app.route('/calculate',methods=['POST'])
def get_msg():
    Session = sessionmaker(bind=engine)
    ses = Session()
    re = ses.query(Calculate).join(User).order_by(Calculate.created.desc())[:10]
    msgs = get_by_list(re)
    return jsonify(msgs)

# アップロード機能
@app.route('/upload', methods=['POST'])
def upload():
        parsonal_data={}
        # ↓　このif節は・・・
        # 申請の年・月・施術者名・施術所名・登録記号番号を入力する
        # ダイアログを通過した場合は、dialog_flg＝True ＞＞よってif節以下は実行されない
        # 通過していない場合（ファイルのアップデートの時）には　dialog_flg = NoneもしくはFalse
        # ＞＞よってif節以下は実行される
        flg1 = request.form.get('dialogFlg')
        #app.logger.info('flg1={}'.format(flg))
        if flg1 =='False':
            
            #https://blog.imind.jp/entry/2020/01/25/032249
            #を参照 拡張子チェック機能１↓ (Excelファイルの拡張子であることを確認するための仕込み段階１)
            ALLOWED_EXTENSIONS = ['.xlsx']
            # ↓ 参照元のサンプルコードでは以下の様だったが、flask.がエラーとなったため削除
            #if 'file' not in flask.request.files:
            if 'file' not in request.files:
                parsonal_data['failed_msg']='読み込めないファイル形式です　アップロード失敗'
                return jsonify(parsonal_data)

            # fileの取得（FileStorage型で取れる）
            # https://tedboy.github.io/flask/generated/generated/werkzeug.FileStorage.html
            
            # ↓ 参照元のサンプルコードでは以下の様だったが、flask.がエラーとなったため削除
            #fs = flask.request.files['file']
            fs = request.files['file']

            # 下記のような情報がFileStorageからは取れる⇒デバックコンソールに表示される仕組みにしてある
            #app.logger.info('file_name={}'.format(fs.filename))
            #app.logger.info('content_type={} content_length={}, mimetype={}, mimetype_params={}'.format(
                #fs.content_type, fs.content_length, fs.mimetype, fs.mimetype_params))
            
            #拡張子チェック機能２↓(Excelファイルの拡張子であることを確認するための仕込み段階２)
            suffix = pathlib.Path(fs.filename).suffix
            #拡張子チェック機能３↓↓(Excelファイルの拡張子であることを確認する段階)
            if not suffix in ALLOWED_EXTENSIONS:
                parsonal_data['failed_msg']="保存できないファイル形式です {}".format(suffix)
                return jsonify(parsonal_data)
            else:
                # ファイルを保存
                fs.save(fs.filename)
                # ↓以下はエクセルを読み込んで、データベースに登録する段取り
            
            path = pathlib.Path("./")    #相対パス指定
            for pass_obj in path.iterdir():
                if pass_obj.match("*.xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
                    """ df=None """
                    df = pd.read_excel(pass_obj,sheet_name = None,header=None,index_col=None)
                    
                    for dfsh in df: ###各シートから読み込んだdataframeのインデックスとヘッダーを番号振りなおし
                        ###Dataframe
                        dfdic=df[dfsh]
                        dfdic.reset_index(drop=True, inplace=True)
                        shp=dfdic.shape
                        dfdic.index=range(1,shp[0]+1)
                        dfdic.columns=range(1,shp[1]+1)
                        global df_new #グローバル変数に値を入れられるようにする
                        # ↑これでdialogから戻ってきても、df_newの値は保持される
                        #pprint.pprint('pre df_new={}'.format(df_new)) 
                        df_new[dfsh]=dfdic ###これで得られたdf_newは、各シート名を[キー]；dataframeを【値】とする辞書
                    
                    # ↓ アップロードされたファイルを、情報を読み取った後に削除
                    # 参考　https://www.atmarkit.co.jp/ait/articles/1910/29/news019_2.html
                    # pathlibライブラリを用いたテクニック。
                    pass_obj.unlink()

        #　↓　変数condDictに、検索条件の辞書を込める
        condDict = get_search_condition()
        #app.logger.info('condDict={}'.format(condDict))
        #　↓　変数sC2cAdicに、辞書を込める
        # ’キー’は'seardhテーブル'の「属性」の文字列：
        # ’値’は’calculateテーブル’の「属性」の文字列
        sC2cAdic = get_dic_schCond2calAttr()
        #　↓　 year_month Dialogから送られてきた変数を、読み込む
        # どういうわけか、ajax通信で送られてきたものは、すべてstring型になってしまうらしい
        year_f = request.form.get('year_fixed')
        month_f = request.form.get('month_fixed')
        therapistName_f = request.form.get('therapistName_fixed')
        treatmentHosName_f = request.form.get('treatmentHosName_fixed')
        registerNo_Str_f = request.form.get('registerNo_Str_fixed')
        
        #app.logger.info('year_month={}'.format(year+'アンド'+month))
        wsh_id_4calc = 1 # calculateテーブルに乗せるデータのidをリセット
        wsh_id_4err = 1 # error_msgテーブルに乗せるデータのidをリセット
        for cD in condDict:
            for dfN_Key in df_new:
                df_value=df_new[dfN_Key]
                ### ↓　DataFrameがある大きさを越えないと、読み込まないようにしておく（はorマ　の申請用紙以外のDataFrameを読み込まない）
                if df_value.shape[0] >= 122 and df_value.shape[1] >= 71 : 
                    if df_value.loc[get_cellno_2list(cD['acupOrMass_Cell'])[0],\
                        get_cellno_2list(cD['acupOrMass_Cell'])[1]] == cD['acupOrMass_Condition']:
                        #　↓　変数d_dicは辞書。後に一気にcalculateテーブルを更新するためのデータを入れとく
                        d_dic={}
                        d_dic['sheetName'] =dfN_Key # シート名を入れておく
                        d_dic['title_AcupOrMass'] =cD['title_AcupOrMass']# はきorマ　を入れておく
                        # ↓「年」が入力されているセル3か所の値がnanではないときに、以下の処理を行う
                        #nanだったときには'False'
                        if not pd.isnull(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                            get_cellno_2list(cD['yearTop_Cell'])[1]] ) and \
                            not pd.isnull(df_value.loc[get_cellno_2list(cD['year1st_Cell'])[0],\
                            get_cellno_2list(cD['year1st_Cell'])[1]]) and\
                            not pd.isnull(df_value.loc[get_cellno_2list(cD['yearLast_Cell'])[0],\
                            get_cellno_2list(cD['yearLast_Cell'])[1]]):

                            # ↓「年」が入力されているセル3か所の、値が一致して、なおかつ　0ではないときに
                            # 　d_dic辞書に　year_Strをキーとして、yearTop_Cellの値を込めておく
                            if int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                get_cellno_2list(cD['yearTop_Cell'])[1]] )) == \
                                int(float(df_value.loc[get_cellno_2list(cD['year1st_Cell'])[0],\
                                get_cellno_2list(cD['year1st_Cell'])[1]]))  and\
                                int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                get_cellno_2list(cD['yearTop_Cell'])[1]]))  == \
                                int(float(df_value.loc[get_cellno_2list(cD['yearLast_Cell'])[0],\
                                get_cellno_2list(cD['yearLast_Cell'])[1]]))  and\
                                int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                get_cellno_2list(cD['yearTop_Cell'])[1]]))  != 0:
                            # ↓「年」がyear month dialogで確認した数字とあっていなければ'False'を入力        
                                if flg1 !='False':
                                    #app.logger.info('year_f={}'.format(year_f))
                                    #app.logger.info('year={}'.format(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    #get_cellno_2list(cD['yearTop_Cell'])[1]] ))
                                    if int(year_f) ==\
                                    int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]])) :
                                        d_dic['year_Str'] =\
                                        str(int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                        get_cellno_2list(cD['yearTop_Cell'])[1]] )))
                                    else:
                                        d_dic['year_Str'] ='False'
                                else:
                                    """ s=df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]]
                                    app.logger.info('s.type={}'.format(type(s))) """
                                    d_dic['year_Str'] =\
                                    str(int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]] )))
                            else:
                                d_dic['year_Str'] ='False'
                        else:
                            d_dic['year_Str'] ='False'
                        # ↓「月」が入力されているセル3か所の値がnanではないときに、以下の処理を行う
                        #nanだったときには'False'
                        if not pd.isnull(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                            get_cellno_2list(cD['monthTop_Cell'])[1]] ) and \
                            not pd.isnull(df_value.loc[get_cellno_2list(cD['month1st_Cell'])[0],\
                            get_cellno_2list(cD['month1st_Cell'])[1]]) and\
                            not pd.isnull(df_value.loc[get_cellno_2list(cD['monthLast_Cell'])[0],\
                            get_cellno_2list(cD['monthLast_Cell'])[1]]):    
                            # ↓「月」が入力されているセル3か所の、値が一致して、なおかつ　0ではないときに
                            # 　d_dic辞書に　month_Strをキーとして、monthTop_Cellの値を込めておく
                            if int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                get_cellno_2list(cD['monthTop_Cell'])[1]] )) == \
                                int(float(df_value.loc[get_cellno_2list(cD['month1st_Cell'])[0],\
                                get_cellno_2list(cD['month1st_Cell'])[1]]))  and\
                                int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                get_cellno_2list(cD['monthTop_Cell'])[1]]))  == \
                                int(float(df_value.loc[get_cellno_2list(cD['monthLast_Cell'])[0],\
                                get_cellno_2list(cD['monthLast_Cell'])[1]]))  and\
                                int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                get_cellno_2list(cD['monthTop_Cell'])[1]]))  != 0:
                            # ↓「月」がyear month dialogで確認した数字とあっていなければ'False'を入力    
                                if flg1 !='False':
                                    if int(month_f) ==\
                                    int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                    get_cellno_2list(cD['monthTop_Cell'])[1]] )):
                                        d_dic['month_Str'] =\
                                        str(int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                        get_cellno_2list(cD['monthTop_Cell'])[1]] )))
                                    else:
                                        d_dic['month_Str'] ='False'
                                #str(int(float・・・とややこしい処理をしているのは、1⃣pandasで
                                #読み込む際に、数値を勝手に「浮動小数点：float」で読み込んで
                                # dataframe化されるがある（例「2」のはずが「2.0」と読み込む）
                                # そのために、2⃣そのデータがstrによって文字列化されてしまうと、
                                # (例「'2.0'」)3⃣それをさらに、intで整数化しようとするとエラーが出る
                                #参考⇒https://qiita.com/ringCurrent/items/1df058bb203374a4b294
                                #これらを回避するために、float関数を用いる


                                else:
                                    d_dic['month_Str'] =\
                                    str(int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                    get_cellno_2list(cD['monthTop_Cell'])[1]] )))
                            else:
                                d_dic['month_Str'] ='False'
                        else:
                            d_dic['month_Str'] ='False'

                        for sC,cA in sC2cAdic.items():
                        # 例）名前　欄が空白でなければ{'name_Cell':名前}／空白ならば{'name_Cell':'False'}
                            
                            # これから判定しなければならないdataframeの各セルのデータを以下の
                            # cellV1に予め込めておく
                            cellV1=df_value.loc[get_cellno_2list(cD[sC])[0],\
                                get_cellno_2list(cD[sC])[1]]
                            # pandasで取得したdataFrameのなかで、欠損値である'nan'は扱いが難しく、
                            # if文で判定するためには、 if cA=='nan' や　if cA==str('nan')では
                            # 判定してくれない（しかも構文エラーにならないので、ややこしい）
                            #　判定するためにはif pd.isnull(?):とする
                            # （欠損値nanならば’true’／pdは　import pandas as pd　より）
                            # 参照⇒https://kagglenote.com/misc/pandas_nan_judge/
                            if not pd.isnull(cellV1) :

                                
                                # ↓　もしも、更新先のテーブルの「属性」に'insurerNo_Str'(保険者番号)という文字列
                                # が含まれていたら、'insurerNoLast_Cell'と'insurerNo_CellStep'を駆使して
                                # 保険者番号を抽出し、'insurerNo_Str'をキーとして
                                # 文字列として入れておく
                                if 'insurerNo_Str' in cA:
                                    number = ''
                                    for n in range(0,8,1):
                                        
                                        ###  ↓　DataFrameの場合、値が入っていない場合は’nan’
                                        #　判定はpd.isnull()
                                        if pd.isnull(df_value.loc[get_cellno_2list(cD[sC])[0],\
                                        get_cellno_2list(cD[sC])[1]+n*cD['insurerNo_CellStep']]):
                                            jj = ''
                                        #　↑　これによって、法別番号（保険者番号の上2桁）が「なし」
                                        # の場合もOK
                                        # ↓保険者番号の1マスに、整数だけでなく、小数を含んだ数字がはいっているかもしれないので、
                                        # str(int(float(によって、無理やり整数化と文字列化をする。
                                        else:
                                            try:
                                                jj = str(int(float(df_value.loc[get_cellno_2list(cD[sC])[0],\
                                                get_cellno_2list(cD[sC])[1]+n*cD['insurerNo_CellStep']] )))
                                            #　↓　str(int(floatでエラーが出る場合は、「数値に変換できない文字列」が入っている場合
                                            # そういう場合は、numberに'False'を入れて、breakでfor文を
                                            # とっとと抜け出す
                                            except:
                                                number= 'False'
                                                break
                                        number = jj + number
                                    d_dic[cA] = number
                                # ↓　もしも、更新先のテーブルの「属性」が'therapistName'(施術者名)だったら
                                # なおかつyear month dialogで確認した施術者名と違っていたら'False'が入る
                                elif 'therapistName' in cA and flg1 !='False' and\
                                    therapistName_f != cellV1 :
                                        d_dic[cA] = 'False'
                                    
                                # ↓　もしも、更新先のテーブルの「属性」が'treatmentHosName'(施術所名)だったら
                                # なおかつyear month dialogで確認した施術所名と違っていたら'False'が入る
                                elif 'treatmentHosName' in cA and flg1 !='False' and\
                                    treatmentHosName_f != cellV1 :
                                        d_dic[cA] = 'False'
                                # ↓　もしも、更新先のテーブルの「属性」が'registerNo_Str'(登録記号番号)だったら
                                # なおかつyear month dialogで確認した登録記号番号と違っていたら'False'が入る
                                elif 'registerNo_Str' in cA and flg1 !='False' and\
                                    registerNo_Str_f != cellV1 :
                                        d_dic[cA] = 'False'
                                
                                # ↓　もしも、更新先のテーブルの「属性」が'amount_Str'(合計額)もしくは
                                # 'copayment_Str'(一部負担金額)もしくは'billingAmount_Str'(請求額)であり、
                                # なおかつyear_month Dialogを開き終わった後だったら・・・
                                # cellV1の値を文字列化したものをd_dic[cA] に入れてみようとする（try文）
                                # エラーが出る（cellV1が文字、もしくはnanだったら）exceptに飛んで、'False'が入る。
                                # うまくいったとしても、cellV1が文字列'0'だったら'False'が入る。
                                # for文から抜け出す　参考⇒https://note.nkmk.me/python-break-nested-loops/
                                # カッコと論理演算子　参考⇒https://dot-blog.jp/news/python-boolean-operations-bool/
                                elif 'amount_Str' in cA or 'copayment_Str' in cA \
                                    or 'billingAmount_Str' in cA :
                                    try:
                                        d_dic[cA] = str(int(float(cellV1)))
                                        if str(int(float(cellV1)))=='0':
                                            d_dic[cA] = 'False'  
                                    except:
                                        d_dic[cA] = 'False'   

                                # 上記以外ならば、素直にセルの値が入る。
                                else:
                                    d_dic[cA] = cellV1
                            else:
                                d_dic[cA] = 'False'
                            """ if cA=='name':
                                app.logger.info('d_dic[name]={}'.format(d_dic[cA]))     """
                        # app.logger.info('d_dic[insurerNo_Str][0:4]={}'.format(d_dic['insurerNo_Str'][0:4]))
                        define_soukatsu1Desti(d_dic)
                        Session = sessionmaker(bind=engine)
                        ses = Session()
                        # ↓ valFalに一つでも'False'文字列が入っていれば、
                        # 'error_msgテーブル'に更新され、
                        # 'False'文字列が入っていなければ、'calculateテーブル'に更新される
                        # ↓　for文のbreakやelseの使い方は 
                        # 右を参照　https://python.civic-apps.com/else-loop/
                        for valFal in d_dic.values():
                            if  valFal=='False'or pd.isnull(valFal) or valFal=='NotFound' or valFal=='0' or valFal=='00000000':
                                d_dic['id'] =wsh_id_4err
                                wsh_id_4err += 1
                                upD_obj = ErrorMsg()
                                upD_obj.update_dict(d_dic)
                                break
                        else:
                            d_dic['id'] =wsh_id_4calc
                            # year_month Dialogにて、年・月・施術者名などを確認するため、
                            # いずれの項目にもFalseがない、一番目のレコードの施術管理者名や
                            #　登録記号番号、施術署名をparsonal_dataにぶっこんで
                            # jsonifyしてreturnで返す
                            if wsh_id_4calc == 1 and flg1 =='False':
                                parsonal_data['therapistName']=d_dic['therapistName']
                                parsonal_data['treatmentHosName']=d_dic['treatmentHosName']
                                parsonal_data['registerNo_Str']=d_dic['registerNo_Str']
                                parsonal_data['year_Int'] =int(d_dic['year_Str'])
                                parsonal_data['month_Int'] =int(d_dic['month_Str'])
                                # ↓ ErrorMsgのレコードを念のためすべて消しておく(先に読み込まれていると、あとで上書きが面倒)
                                ses.query(ErrorMsg).delete()
                                ses.commit()
                                ses.close()
                                
                                return jsonify(parsonal_data)
                            wsh_id_4calc += 1
                            upD_obj = Calculate()
                            upD_obj.update_dict(d_dic)
                        ses.add(upD_obj)
                        ses.commit()
                        ses.close()
        
        parsonal_data['process_msg']='総括票　作成中・・・' 
        
        #　↓　ブックの複製　参照⇒https://neko-py.com/python-excel-write-book
        wb = openpyxl.load_workbook(filename='soukatsuTemp.xlsx')
        #　↓　日付や時間の取得　参照⇒https://www.sejuku.net/blog/23606
        # しかし、上記のとおりに、date = datetime.datetime.now()　と書くとエラー
        date = datetime.now()
        # 2桁表示のゼロパディングは　参照⇒https://note.nkmk.me/python-zero-padding/
        sortInsList=sort_insureName_4Sokatsu1()
        template_sheet = wb['総括票（Ⅰ）(ひな形　禁削除)']
        # ↓　総括表１の送付先（soukatsu1Desti）だけを、重複なくリスト化したものがsoukatsu1Desti_List
        soukatsu1Desti_List = soukatsu1Desti_List_set(sortInsList)

        re = ses.query(Calculate).all()
        loadD_obj = get_by_list(re)
        ses.query(Calculate).delete()# 復帰
        ses.commit()# 復帰
        ses.close()# 復帰
        re = ses.query(ErrorMsg).all()
        ErrD_obj = get_by_list(re)
        ses.query(ErrorMsg).delete() #復帰
        ses.commit()#復帰
        ses.close()# 復帰

        for desti in soukatsu1Desti_List:
            dicDesti_insur= get_soukatsu1Desti_insur_dic(sortInsList,soukatsu1Desti_List) 
            #app.logger.info('desti={}'.format(desti))  
            #app.logger.info('count1={}'.format(counter1))                     
            #app.logger.info('count1int={}'.format((int(counter1 / 7 - 0.1)+1)))
            # ↓同じ総括表１の行き先（soukatsu1Desti）に、どれだけの保険者の数がはいるか？
            # を、変数yyに込める
            yy=int(len(dicDesti_insur[desti]) / 7 - 0.1)+1
            for x in range(yy):     
                    target_sheet = wb.copy_worksheet(template_sheet)
                    #app.logger.info('template_sheet={}'.format(template_sheet.sheet_properties.tabColor))
                    # ↓　複製したシートのタブの色を、色なしにする
                    target_sheet.sheet_properties.tabColor =None
                    if x >= 1:
                        target_sheet.title = '総括票（Ⅰ）('+ desti +'）'+str(x+1)+'枚目'
                    else: 
                        target_sheet.title = '総括票（Ⅰ）('+ desti +'）'
                    # ↓ 複製したシート(総括票（Ⅰ）)に、それぞれの数値を入れる
                    target_sheet.cell(14, 26).value = registerNo_Str_f #登録記号番号
                    target_sheet.cell(17, 26).value = therapistName_f #施術管理者
                    target_sheet.cell(20, 26).value = treatmentHosName_f #施術所名
                    target_sheet.cell(6, 10).value = year_f #年
                    target_sheet.cell(6, 21).value = month_f #月
                    # ↓ 複製したシートが複数にわたり、x枚目がその最後の時は、
                    # 保険者名が7段目までいかずに、途中(y段目)で終わるような仕掛け
                    if x == yy-1:
                        y = len(dicDesti_insur[desti]) - 7*(x)
                    # ↓ 複製したシートが複数にわたり、x枚目がその途中の時は、
                    # 7段目までフルに入力する
                    elif x < (yy-1):
                        y = 7
                    #app.logger.info('y={}'.format(y))
                    #app.logger.info('target_sheet.title={}'.format(target_sheet.title))
                    
                    # ↓ 複製したシート（総括表１）に、保険者名＋改行＋（はりきゅうorマッサージ）
                    # を、上から順にｖ番目まで入れていく
                    # （数列v+7*xを用いて、コピーがx枚目のときは、1枚目の続きの保険者が入るようにしてある）
                    v=0
                    for v in range(y):
                        listv = dicDesti_insur[desti]
                        target_sheet.cell(32 + 6 * v, 2).value = listv[v+7*x][0]+'\n'+'('+listv[v+7*x][1]+')'
                        
                        for loadD in loadD_obj:
                            if loadD['kanji_Insurer_Name'] == listv[v+7*x][0] \
                                and loadD['title_AcupOrMass'] == listv[v+7*x][1]:
                                if loadD['relationship'] == '本人':
                                    target_sheet_cell1=target_sheet.cell(34 + 6 * v, 15)#本人の件数を入れるセル
                                    kensuu_insert(target_sheet_cell1)
                                    target_sheet_cell2=target_sheet.cell(34 + 6 * v, 21)#本人の費用額を入れるセル
                                    loadDInt = int(float(loadD['amount_Str']))#本人の費用額
                                    kingaku_insert(loadDInt,target_sheet_cell2)
                                else:
                                    target_sheet_cell1=target_sheet.cell(34 + 6 * v, 33)#家族の件数を入れるセル
                                    kensuu_insert(target_sheet_cell1)
                                    target_sheet_cell2=target_sheet.cell(34 + 6 * v, 39)#家族の費用額を入れるセル
                                    loadDInt = int(float(loadD['amount_Str']))#家族の費用額
                                    kingaku_insert(loadDInt,target_sheet_cell2)
        template_sheet = wb['総括票（Ⅱ）(ひな形　禁削除)']            
        for insurL in sortInsList:
            target_sheet = wb.copy_worksheet(template_sheet)
            target_sheet.sheet_properties.tabColor =None
            target_sheet.title = '総括票（Ⅱ）' + insurL[3] + '(' + insurL[2][0] +  ')'
            target_sheet.cell(18, 26).value = registerNo_Str_f #登録記号番号
            target_sheet.cell(21, 26).value = therapistName_f #施術管理者
            target_sheet.cell(24, 26).value = treatmentHosName_f #施術所名
            target_sheet.cell(6, 10).value = year_f #年
            target_sheet.cell(6, 21).value = month_f #月
            target_sheet.cell(14, 15).value = insurL[3]+'\n'+'('+insurL[2] +')' #保険者名＋はorマ
            for loadD in loadD_obj:
                if loadD['kanji_Insurer_Name'] == insurL[3] \
                    and loadD['title_AcupOrMass'] == insurL[2]:
                    if loadD['relationship'] == '本人':
                        target_sheet_cell1=target_sheet.cell(35, 13)#本人の件数を入れるセル
                        kensuu_insert(target_sheet_cell1)
                        target_sheet_cell2=target_sheet.cell(35, 24)#本人の費用額を入れるセル
                        loadDInt = int(float(loadD['amount_Str']))#本人の費用額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(35, 39)#本人の一部負担金額を入れるセル
                        loadDInt = int(float(loadD['copayment_Str']))#本人の一部負担金額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(35, 54)#本人の請求額を入れるセル
                        loadDInt = int(float(loadD['billingAmount_Str']))#本人の請求額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                    else:
                        target_sheet_cell1=target_sheet.cell(41, 13)#家族の件数を入れるセル
                        kensuu_insert(target_sheet_cell1)
                        target_sheet_cell2=target_sheet.cell(41, 24)#家族の費用額を入れるセル
                        loadDInt = int(float(loadD['amount_Str']))#家族の費用額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(41, 39)#家族の一部負担金額を入れるセル
                        loadDInt = int(float(loadD['copayment_Str']))#家族の一部負担金額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(41, 54)#家族の請求額を入れるセル
                        loadDInt = int(float(loadD['billingAmount_Str']))#家族の請求額
                        kingaku_insert(loadDInt,target_sheet_cell2)
        
        template_sheet = wb['総括表　新潟県師会用（禁削除）']            
        target_sheet = wb.copy_worksheet(template_sheet)
        target_sheet.sheet_properties.tabColor =None
        target_sheet.title = '総括表　新潟県師会用('+str(date.month).zfill(2)+'月' +str(date.day).zfill(2) +'日'+ str(date.hour).zfill(2)+'時' + str(date.minute).zfill(2) +'分 作成'+')' 
        target_sheet.cell(2, 2).value = year_f #年
        target_sheet.cell(2, 4).value = month_f #月
        target_sheet.cell(3, 5).value = therapistName_f #施術管理者
        for loadD in loadD_obj:
            #app.logger.info('insurerNo_Str={}'.format(loadD['kanji_Insurer_Name']+loadD['insurerNo_Str'][0:1]))
            if '協会' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(7, 2)#協会けんぽのはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(7, 4)#協会けんぽのはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(7, 6)#協会けんぽのマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(7, 7)#協会けんぽのマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif '共済' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(9, 2)#共済のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(9, 4)#共済のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(9, 6)#共済のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(9, 7)#共済のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif '国民健康保険組合' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(10, 2)#国保組合のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(10, 4)#国保組合のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(10, 6)#国保組合のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(10, 7)#国保組合のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            #保険者番号が6桁もしくは、山形県のように5桁の場合、
            # あるいは67から始まる退職者医療の場合⇒国保へ分類される
            #なお、国保組合も、6桁～5桁だが、前elif節のところで
            #スクリーニングしてあるので心配いらない
            elif len(loadD['insurerNo_Str']) == 6 or len(loadD['insurerNo_Str']) == 5 or loadD['insurerNo_Str'][0:2] == '67' :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(12, 2)#国保のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(12, 4)#国保のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(12, 6)#国保のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(12, 7)#国保のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif loadD['insurerNo_Str'][0:2] == '39':
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(13, 2)#後期高齢のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(13, 4)#後期高齢のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(13, 6)#後期高齢のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(13, 7)#後期高齢のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
        wb.remove(wb['総括表　新潟県師会用（禁削除）'])
        wb.remove(wb['総括票（Ⅰ）(ひな形　禁削除)'])
        wb.remove(wb['総括票（Ⅱ）(ひな形　禁削除)'])
        koukikourei_No_Sort(loadD_obj,wb)
        
        parsonal_data['alert_data'] = error_Msg_Sheet(ErrD_obj,wb)
        #app.logger.info('parsonal_data[alert_data] ={}'.format(parsonal_data['alert_data'] ))  
        #app.logger.info('dicDesti_insur={}'.format(dicDesti_insur))

        # ↓　.xlsxファイルが作成された年月日と時刻を、日本時間で取得（サーバーのおかれている国によって、時刻が変動しないように）
        # 参考⇒https://qiita.com/keisuke0508/items/df2594770d63bf124ccd
        now = datetime.now(pytz.timezone('Asia/Tokyo'))
        # ↓　作成された.xlsxファイルに、作成年月日でファイル名を命名する
        # ↓　時刻の2桁表示（ゼロ埋め）は.zfill()で行う
        # 参考⇒https://note.nkmk.me/python-zero-padding/
        try:
            dLFileName='総括票 令和'+year_f+'年'+month_f+'月分　'+str(now.month).zfill(2)+'月' +str(now.day).zfill(2) +'日'+ str(now.hour).zfill(2)+'時' + str(now.minute).zfill(2) +'分'+str(now.second).zfill(2) +'秒 作成'+ '.xlsx'
        except:
            dLFileName='すべてのシートが読み込み不可'+ '.xlsx'
        
        wb.save(dLFileName)
        wb.close() 
        parsonal_data['dLFile']=dLFileName

        return  jsonify(parsonal_data)
        # '総括票 令和'+year_f+'年'+month_f+'月分　'+str(date.month).zfill(2)+'月' +str(date.day).zfill(2) +'日'+ str(date.hour).zfill(2)+'時' + str(date.minute).zfill(2) +'分'+str(date.second).zfill(2) +'秒 作成'+ '.xlsx')

@app.route('/download', methods=['POST'])
def download(): 
    #　↓　作成されたファイル名をindexのload_file()から送られてきた
    # formDataから読み取り、ファイルをsend_fileによってindexに送って
    # ダウンロードさせる
    # 参照⇒https://qiita.com/5zm/items/760000cf63b176be544c#2-%E6%96%B9%E6%B3%95%EF%BC%91send_file%E3%82%92%E5%88%A9%E7%94%A8%E3%81%99%E3%82%8B
    fName= request.form.get('filename')  
    #app.logger.info('fName={}'.format(fName)) 
    download_file_name=fName
    download_file = fName
    XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    sendingFile=send_file(download_file, as_attachment=True,\
              attachment_filename=download_file_name,\
              mimetype=XLSX_MIMETYPE)

    return sendingFile

@app.route('/dLfileDel', methods=['GET'])
def dlf_delete(): 
    path = pathlib.Path("./") 
    for pass_obj in path.iterdir():
            if pass_obj.match("*.xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
                # ↓ 作成された総括票ファイルを、ダウンロード後に削除
                # 参考　https://www.atmarkit.co.jp/ait/articles/1910/29/news019_2.html
                # pathlibライブラリを用いたテクニック。
                pass_obj.unlink()
    return '総括票を作成しました。ダウンロードして「名前を付けて保存」してください。'
if __name__=='__main__':
    app.debug = True

    # ↓　サーバーにデプロイして公開するためのもの
    app.run(host='0.0.0.0')
    # ↓　ローカルで用いるためのもの
    #app.run(host='localhost')