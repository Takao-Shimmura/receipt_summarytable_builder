#!python3.9.1
from flask import Flask, session
    #redirect, jsonify, current_app, g
#import psycopg2
import openpyxl # 外部ライブラリ　pip install openpyxl
import sqlite3
#import json
from datetime import datetime
import pprint

import os
import pathlib

from sqlalchemy import create_engine, Column, Integer, String, \
    Text, DateTime, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy.orm.exc import NoResultFound


#　↓　herokuのpostgreSQL接続用URI 
# ※ただし、割り当てられたURIそのままでは接続エラー
#　「postgres://・・・」から「postgresql://・・・」に変更しなければ解消されない
#参考（heroku公式リファレンス）⇒Why is SQLAlchemy 1.4.x not connecting to Heroku Postgres? - Heroku Help
engine = create_engine('postgresql://qrnkdpytaiifps:7b728dc1e568e2d1c1ab80c919e17d10c7f41f8d853c8e5989d907c978bf8d8c@ec2-34-250-16-127.eu-west-1.compute.amazonaws.com:5432/d77prcb2vt5pne')

#　↓　ローカルのSQLite接続用パス 
###engine = create_engine('sqlite:///sample.sqlite3')



# base model
Base=declarative_base()
#model class

                
class Search_condition(Base):
    __tablename__='search'

    id=Column(Integer,primary_key=True)
    
    # ※県単の書類の場合、「読み取らないセル」は、データベースに文字列'pass'を入れておく
    condition_Title = Column(String(255))
    #「condition_Title」はデータベースを閲覧するときの表題
    # として設けたもので、Flaskでは使用しない
    # ※例外的に県単の書類かどうかの判定に用いる⇒「県単医療費助成申請書」という文字列
    title_AcupOrMass= Column(String(255))
    #「title_AcupOrMass」は「はりきゅう」／「マッサージ」のいずれかが入る
    acupOrMass_Condition = Column(String(255))
    #「acupOrMass_Condition」は「（はり・きゅう用）」／
    # 「（あんま・マッサージ用）」のいずれかが入る（新書式になって変更もありうる）
    
    # ↓↓　これ以降の項目に入る値は、「対象となるセル」の「行数」と「列数」を
    #'?_?'という文字列にしたもの

    acupOrMass_Cell = Column(String(255))
    # ↑「対象セル」⇒（はり・きゅう用）もしくは（あんま・マッサージ用）が入力されているセル
    # ↑※例外的に、「県単医療費助成申請書」が入っているセルにも用いる
    insurerNoLast_Cell = Column(String(255))
    # ↑ 「対象セル」⇒保険者番号の末尾の番号が入力されているセル
    insurerNo_CellStep=Column(Integer)
    # ↑保険者番号の末尾の番号 から左に何セルずつずれていくと、
    #次の番号になるか？　を数字で表したもの　マイナスの値になる
    insuraCodeNo_Cell = Column(String(255))
    # ↑ 「対象セル」⇒被保険者証等の記号番号が入力されているセル
    name_Cell = Column(String(255))
    # ↑ 「対象セル」⇒施術を受けた者の氏名が入力されているセル
    nameKana_Cell = Column(String(255))
    # ↑ 「対象セル」⇒施術を受けた者の氏名の読み仮名（カタカナ）が入力されているセル
    amount_Cell = Column(String(255))
    # ↑ 「対象セル」⇒合計　のセル
    copayment_Cell = Column(String(255))
    # ↑ 「対象セル」⇒一部負担金　のセル
    billingAmount_Cell = Column(String(255))
    # ↑ 「対象セル」⇒請求額　のセル
    relationship_Cell = Column(String(255))
    # ↑ 「対象セル」⇒続柄　のセル
    #<2021 7月分より削除>insuredName_Cell = Column(String(255))
    # ↑ 「対象セル」⇒被保険者（申請書の下方の「申請欄」）のセル
    therapistName_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の「施術者名」のセル
    treatmentHosName_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の「施術所名」のセル
    registerNo_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の「登録記号・番号」のセル
    yearTop_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の一番上の行の、「年」のセル
    year1st_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の施術期間の、開始日の「年」のセル
    yearLast_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の施術期間の、終了日の「年」のセル
    # ※県単の場合は、yearTop/1st/Lastいずれも同じセルにしておく

    monthTop_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の一番上の行の、「月」のセル
    month1st_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の施術期間の、開始日の「月」のセル
    monthLast_Cell= Column(String(255))
    # ↑ 「対象セル」⇒「申請欄」の施術期間の、終了日の「月」のセル
    # ※県単の場合は、monthTop/1st/Lastいずれも同じセルにしておく

    # get Dict data
    def to_dict(self):
        return{
            'id':int(self.id),
            'condition_Title':str(self.condition_Title),#　実はこの項目はデータベース上の　レコードのタイトル　であるのでほぼ使用しない
            'title_AcupOrMass':str(self.title_AcupOrMass),
            'acupOrMass_Condition':str(self.acupOrMass_Condition),
            'acupOrMass_Cell':str(self.acupOrMass_Cell),
            'insurerNoLast_Cell':str(self.insurerNoLast_Cell),
            'insurerNo_CellStep':int(self.insurerNo_CellStep),
            'insuraCodeNo_Cell':str(self.insuraCodeNo_Cell),
            'name_Cell':str(self.name_Cell),
            'nameKana_Cell':str(self.nameKana_Cell),
            'amount_Cell':str(self.amount_Cell),  
            'copayment_Cell':str(self.copayment_Cell),  
            'billingAmount_Cell':str(self.billingAmount_Cell),  
            'relationship_Cell':str(self.relationship_Cell),  
            #<2021 7月分より>'insuredName_Cell':str(self.insuredName_Cell), 
            'therapistName_Cell':str(self.therapistName_Cell),
            'treatmentHosName_Cell':str(self.treatmentHosName_Cell),
            'registerNo_Cell':str(self.registerNo_Cell),
            'yearTop_Cell':str(self.yearTop_Cell),
            'year1st_Cell':str(self.year1st_Cell),
            'yearLast_Cell':str(self.yearLast_Cell),
            'monthTop_Cell':str(self.monthTop_Cell),
            'month1st_Cell':str(self.month1st_Cell),
            'monthLast_Cell':str(self.monthLast_Cell),
                }
class InsurerData(Base):
    __tablename__='insurerdata'

    id=Column(Integer,primary_key=True)
    kanji_Insurer_Name= Column(String(255))
    kana_Insurer_Name= Column(String(255))
    insurer_No_Str= Column(String(255))
    soukatsu1Desti= Column(String(255))
    info= Column(String(255))
    info_admin= Column(String(255))
    # get Dict data
    def to_dict(self):
        return{
            'id':int(self.id),
            'kanji_Insurer_Name':str(self.kanji_Insurer_Name),
            'kana_Insurer_Name':str(self.kana_Insurer_Name),
            'insurer_No_Str':str(self.insurer_No_Str),
            'soukatsu1Desti':str(self.soukatsu1Desti), 
            'info':str(self.info), 
            'info_admin':str(self.info_admin), 
        }


def get_dic_schCond2calAttr():
    return{
        'insurerNoLast_Cell':'insurer_No_Str',
        'insuraCodeNo_Cell':'insuraCodeNo_Str',
        'name_Cell':'name',
        'nameKana_Cell':'nameKana',
        'amount_Cell':'amount_Str',  
        'copayment_Cell':'copayment_Str',  
        'billingAmount_Cell':'billingAmount_Str',  
        'relationship_Cell':'relationship',  
        # 'insuredName_Cell':'insuredName', <2021 7月分より>
        'therapistName_Cell':'therapistName',
        'treatmentHosName_Cell':'treatmentHosName',
        'registerNo_Cell':'registerNo_Str', 
        
          
    }
#検索の条件をDBのSearchテーブルから引き出す

def get_search_condition():
    Session = sessionmaker(bind=engine)
    ses = Session()
    re = ses.query(Search_condition).all()
    conditions = get_by_list(re)
    ses.close()# 終わったら必ずセッションを閉じておかないと、SQLalchemy内でのエラーが出る（それでも動作は完遂してくれるが）
    return conditions

#保険者番号一覧をDBのInsurerdataテーブルから引き出す

def get_insurerData_all():
    Session = sessionmaker(bind=engine)
    ses = Session()
    re = ses.query(InsurerData).all()
    insurerData_all = get_by_list(re)
    ses.close()# 終わったら必ずセッションを閉じておかないと、SQLalchemy内でのエラーが出る（それでも動作は完遂してくれるが）
    #print('insurerData_all={}'.format(insurerData_all)) 
    return insurerData_all

# ↓　get_by_list()関数にて個々の検索条件を「辞書」に。
# そして複数の辞書（検索条件）をまとめて、リスト化する。
def get_by_list(arr):
    res = []
    for item in arr:
        res.append(item.to_dict())
    for eachres in res:
        #　↓　保険者番号一覧　での備考（'info'項目）が、
        #空欄だった場合、あえて空文字''を入れておいて、
        #一覧表にNoneと表示されないようにする
        if eachres['info'] == 'None':
            eachres['info']=''
        #print('res={}'.format(res)) 
    return res 

# 辞書化された検索の条件から、'?_?'という文字列を、
# [?,?]というリストに変換する
def get_cellno_2list(cellint):
    li2=cellint.split('_')
    return [int(str) for str in li2]

# nameから、'姓 名'という文字列を、
# [姓,名]というリストに変換する
def get_name_split_space(name1):
    if '  ' in name1: #半角スペース2つの場合
        li2=name1.split('  ')
        return [str for str in li2]
    elif '　　' in name1: #全角スペース2つの場合
        li2=name1.split('　　')
        return [str for str in li2]
    elif '　 ' in name1: #全角スペース＋半角スペースの場合
        li2=name1.split('　 ')
        return [str for str in li2]
    elif ' 　' in name1: #半角スペース＋全角スペースの場合
        li2=name1.split(' 　')
        return [str for str in li2]
    elif '　' in name1: #全角スペース1つの場合
        li2=name1.split('　')
        return [str for str in li2]
    elif ' ' in name1: #半角スペース1つの場合
        li2=name1.split(' ')
        return [str for str in li2]
    

def define_soukatsu1Desti(dic1):
    Session = sessionmaker(bind=engine)
    ses = Session()
    #　↓　try文でd_dic['insurer_No_Str'}の内容がInsurerDataを検索して一致すれば、
    #  dic1['soukatsu1Desti']1に、保険者の宛先を込めることができる。
    # except文の中のif文で、Falseのときと、そうでないときを条件分岐しており、
    # さらなるtry文によって後期高齢者の保険者番号のときに、先頭4桁＋’****’の文字列を検索。
    #　except文で保険者番号がまったく登録されていない場合に分けている。
    # ＞＞＞再考が必要
    #　↓　SQLalchemyの前方一致検索のやり方は
    # 参考　https://mycodingjp.blogspot.com/2019/07/flask-sqlalchemy.html
    #　↓　try exceptの例外処理のやり方は
    # 参考　https://www.atmarkit.co.jp/ait/articles/1909/06/news019.html
    try:
        myinsdata = ses.query(InsurerData).\
        filter(InsurerData.insurer_No_Str==dic1['insurer_No_Str']).one()
        dic1['soukatsu1Desti'] = myinsdata.soukatsu1Desti
        dic1['kana_Insurer_Name'] = myinsdata.kana_Insurer_Name
        dic1['kanji_Insurer_Name'] = myinsdata.kanji_Insurer_Name
    
    except NoResultFound:
        if dic1['insurer_No_Str'] =='False':
            dic1['soukatsu1Desti'] = 'False'
            dic1['kana_Insurer_Name'] = 'False'
            dic1['kanji_Insurer_Name'] = 'False'
        #　↓　協会けんぽ　組合保険　の一番最初の桁が0始まりなので、
        # insurardataに登録されている保険番号と合致しない。そのときは
        # 先頭の0を消去して、再検索してみる
        elif dic1['insurer_No_Str'][0:1] =='0':
            try:
                myinsdata = ses.query(InsurerData).\
                filter(InsurerData.insurer_No_Str==dic1['insurer_No_Str'][1:] ).one()
                dic1['soukatsu1Desti'] = myinsdata.soukatsu1Desti
                dic1['kana_Insurer_Name'] = myinsdata.kana_Insurer_Name
                dic1['kanji_Insurer_Name'] = myinsdata.kanji_Insurer_Name
            except:
                dic1['soukatsu1Desti'] = 'NotFound'
                dic1['kana_Insurer_Name'] = 'NotFound'
                dic1['kanji_Insurer_Name'] = 'NotFound'
        #　↓　国保の退職者医療　の一番最初の桁が67始まりなので、
        # insurardataに登録されている保険番号と合致しない。そのときは
        # 先頭の67を消去して、再検索してみる
        elif dic1['insurer_No_Str'][0:2] =='67':
            try:
                myinsdata = ses.query(InsurerData).\
                filter(InsurerData.insurer_No_Str==dic1['insurer_No_Str'][2:] ).one()
                dic1['soukatsu1Desti'] = myinsdata.soukatsu1Desti
                dic1['kana_Insurer_Name'] = myinsdata.kana_Insurer_Name
                dic1['kanji_Insurer_Name'] = myinsdata.kanji_Insurer_Name
            except:
                #　先頭の67を消去して、再検索してみても市町村の国保と合致しない場合
                # 山形県のように保険者番号が５桁の場合もあるので、
                # 下5桁（[3:] ）で検索してみる
                try:
                    myinsdata = ses.query(InsurerData).\
                    filter(InsurerData.insurer_No_Str==dic1['insurer_No_Str'][3:] ).one()
                    dic1['soukatsu1Desti'] = myinsdata.soukatsu1Desti
                    dic1['kana_Insurer_Name'] = myinsdata.kana_Insurer_Name
                    dic1['kanji_Insurer_Name'] = myinsdata.kanji_Insurer_Name
                except:
                    dic1['soukatsu1Desti'] = 'NotFound'
                    dic1['kana_Insurer_Name'] = 'NotFound'
                    dic1['kanji_Insurer_Name'] = 'NotFound'
        #　↓　後期高齢者者医療　の一番最初の４桁は、と都道府県ごとに決まっている。（下4桁は市町村で異なる）
        # insurardataに登録されている保険番号は上4桁のみ　で、下4桁はアスタリスク「****」にしてある。
        # 上4桁を残して、下4桁を「****」に変えて、再検索してみる
        else:
            try:
                myinsdata = ses.query(InsurerData).filter(InsurerData.insurer_No_Str==dic1['insurer_No_Str'][0:4]+'****').one()
                dic1['soukatsu1Desti'] = myinsdata.soukatsu1Desti
                dic1['kana_Insurer_Name'] = myinsdata.kana_Insurer_Name
                dic1['kanji_Insurer_Name'] = myinsdata.kanji_Insurer_Name
            except:
                dic1['soukatsu1Desti'] = 'NotFound'
                dic1['kana_Insurer_Name'] = 'NotFound'
                dic1['kanji_Insurer_Name'] = 'NotFound'
    ses.close()# 終わったら必ずセッションを閉じておかないと、SQLalchemy内でのエラーが出る（それでも動作は完遂してくれるが）
    return dic1
    
# ↓　2重のリストで、重複しないリストを作成　参考⇒https://note.nkmk.me/python-list-unique-duplicate/
def get_unique_list(listInList):
    list1 = []
    """ for x in listInList:
        if x not in list1 and not list1.append(x):
            list1.append(x) """
    # ↑　上記コメントアウトしている部分と
    # ↓　下記のreturnで返している戻り値の「リスト内包表記」は同じ内容。
    #  特筆すべきは、if文中のnot list1.append(x)という表現
    #　この表現は2つの役割がある　
    # ①「x が　list1内に存在しないとき（x not in list1）」に、list1にxを追加（append）する
    # ②　list1にxを追加しつつも、返す値はnoneなので、not list1.append(x)はTrueとなり、
    # if文を成立させる
    # (破壊的リストになるのでpythonでは　＝list1.append(x)　としてもnoneが返される)
    return [x for x in listInList if x not in list1 and not list1.append(x)] 

# リストのリストと、総括表の行き先（soukatsu1Desti）を引数として、
#　総括票の行き先がおなじリストが何個存在するか？をカウントする関数
def get_soukatsu1Desti_count(listInList,soukatsu1Desti):
        counter1 = 0
        #pprint.pprint('listInList={}'.format(listInList)) 
        for x in listInList:
            #pprint.pprint('x={}'.format(x)) 
            #print('soukatsu1Desti={}'.format(soukatsu1Desti)) 
            if soukatsu1Desti in x :
                #print('soukatsu1Desti in if={}'.format(soukatsu1Desti)) 
                counter1 += 1
        return counter1

# ↓リストのリスト(L)と、総括表の行き先（soukatsu1Desti）を引数として、
#　{総括票の行き先1:[[保険者1,'はりきゅう'or'マッサージ'],
#                   [保険者2,'はりきゅう'or'マッサージ']・・・],
#  {総括票の行き先2:[[保険者1,'はりきゅう'or'マッサージ'],
#                   [保険者2,'はりきゅう'or'マッサージ']・・・]}
# というリストinリストin辞書を作成する関数
def get_soukatsu1Desti_insur_dic(listInList,soukatsu1Desti):
    dic1={}
    for key1 in soukatsu1Desti:
        if '△' not in key1:# 総括票Ⅰを作りたくない保険者
            list0=[]
            for list2 in listInList:
                if list2[0] == key1:
                    list1=[]
                    list1.append(list2[3])
                    list1.append(list2[2])
                    if list1 not in list0:
                        list0.append(list1)
                        #print('list0={}'.format(list0)) 
            dic1[key1]=list0
    return dic1

# データベースのCalculateテーブルから
# 保険者の総括票1・2の順番を決めるリストのリストを作成する
# insDestSort1 in insDestSort2 のリスト2重構造にする
def sort_insureName_4Sokatsu1_fromloadD_obj(obj):
    
    calcu_list = obj
    insDestSort2=[]
    
    
    for cL in calcu_list:
        insDestSort1=[]# 保険者の総括票1の順番を決めるリスト
        insDestSort1.append(cL['soukatsu1Desti'])
        insDestSort1.append(cL['kana_Insurer_Name'])
        insDestSort1.append(cL['title_AcupOrMass'])
        insDestSort1.append(cL['kanji_Insurer_Name'])
        #pprint.pprint('insDestSort1={}'.format(insDestSort1)) 
        insDestSort2.append(insDestSort1)
        #pprint.pprint('insDestSort2={}'.format(insDestSort2)) 
    # ↓　2重のリストで、重複しないリストを作成する関数
    insDestSort2=get_unique_list(insDestSort2)
    insDestSort2.sort()
    
    return insDestSort2

# sort_insureName_4Sokatsu1で作ったリストのリストから
# soukatsu1Desti（総括票の行き先）だけを抽出して、
# 重複の無いようなリストを作る関数
# 総括表Ⅰを何枚複製するのか？を決めるために必要。
def soukatsu1Desti_List_set(listInList):
    list1=[]
    for list2 in listInList:
        list1.append(list2[0])
        list1=list(set(list1))
    return list1

def kensuu_insert(target_sheet_cell1):
    #print('target_sheet_cell1.value={}'.format(target_sheet_cell1.value)) 
    if target_sheet_cell1 != None:
        if target_sheet_cell1.value == None:
            target_sheet_cell1.value = 1
        elif target_sheet_cell1.value != None:
            target_sheet_cell1.value = int(target_sheet_cell1.value) + 1
    return

def kingaku_insert(loadDInt,target_sheet_cell2):
    if target_sheet_cell2.value == None:
        target_sheet_cell2.value = int(loadDInt)
    elif target_sheet_cell2.value != None:
        target_sheet_cell2.value = int(target_sheet_cell2.value) + loadDInt
    return

def error_Msg_Sheet(err_obj,errKentan_obj,wb2):
    #pprint.pprint('err_obj={}'.format(err_obj)) 
    alt_data=[]
    if err_obj==[] and errKentan_obj==[]:
        return 'False'
    else:
        if err_obj!=[] :    
        # ちなみに、err_objはE申請書のエラーデータ（辞書inリスト）
        # alt_dataは、indexへの戻り値となるparsonal_dataに込めるエラーメッセージのデータ（辞書inリスト）のこと
            for l4 in err_obj:
                list4=[]
                list4.append('★読み込みができなかったシート：　【'+l4['sheetName']+'】')
                if l4['name']=='False'or l4['name']== '0':
                    list4.append([1,'「療養を受けた者の氏名」の記入漏れ'])
                if l4['nameKana']== 'False' or l4['nameKana']== '0':
                    list4.append([2,'「療養を受けた者の氏名」(フリガナ))の記入漏れ'])
                if l4['insurer_No_Str']== 'False':
                    list4.append([3,'「保険者番号」の記入漏れ'])
                if l4['insuraCodeNo_Str']=='False'or l4['insuraCodeNo_Str']== '0':
                    list4.append([4,'「被保険者証等の記号番号」の記入漏れ'])
                if l4['amount_Str']=='False' :
                    list4.append([5,'「合計」金額の記入漏れ'])
                if l4['copayment_Str']== 'False' :
                    list4.append([6,'「一部負担金」金額の記入漏れ'])
                if l4['billingAmount_Str']== 'False' :
                    list4.append([7,'「請求額」金額の記入漏れ'])
                if l4['relationship']== 'False' or l4['relationship']== '0':
                    list4.append([8,'「続柄」の記入漏れ'])
                # <2021 7月分より>if l4['insuredName']== 'False' or l4['insuredName']== '0':
                    #list4.append([9,'「申請者（被保険者）」氏名の記入漏れ']) 
                if l4['therapistName']== 'False':
                    list4.append([9,'施術管理者の「氏名」の記入漏れもしくは記入ミス'])
                if l4['treatmentHosName']== 'False':
                    list4.append([10,'施術管理者の「名称」の記入漏れもしくは記入ミス'])
                if l4['registerNo_Str']== 'False':
                    list4.append([11,'施術管理者の「登録記号番号」の記入漏れもしくは記入ミス'])
                if l4['year_Str']== 'False':
                    list4.append([12,'申請書冒頭の申請「年」か、施術期間の「年」のいずれかの記入漏れもしくは記入ミス'])
                if l4['month_Str']== 'False':
                    list4.append([13,'申請書冒頭の申請「月」か、施術期間の「月」のいずれかの記入漏れもしくは記入ミス'])
                if l4['kanji_Insurer_Name']== 'NotFound':
                    list4.append([14,'「保険者番号」から保険者が特定できません　保険者番号の記入ミスもしくは、ホームページ管理者による「保険者番号の登録漏れ」です'])
                list4.append(['',''])
                alt_data.append(list4)

        if errKentan_obj!=[]:
            # ちなみに、errKentan_objは県単医療費助成のエラーデータ（辞書inリスト）
            # alt_dataは、indexへの戻り値となるparsonal_dataに込めるエラーメッセージのデータ（辞書inリスト）のこと
            for l7 in errKentan_obj:
                list4=[]
                list4.append('★読み込みができなかったシート：　【'+l7['sheetName']+'】')
                if l7['name']=='False'or l7['name']== '0':
                    list4.append([15,'【県単】「受給者氏名」の記入漏れ'])
                if l7['amount_Str']=='False' :
                    list4.append([16,'【県単】「合計」金額の記入漏れ'])
                if l7['year_Str']== 'False':
                    list4.append([17,'【県単】冒頭の申請「年」、診療月の「年」2か所 のいずれかの記入漏れもしくは記入ミス'])
                if l7['month_Str']== 'False':
                    list4.append([18,'【県単】冒頭の申請「月」、診療月の「月」2か所 のいずれかの記入漏れもしくは記入ミス'])
                list4.append(['',''])
                alt_data.append(list4)
            # ↓　第２引数で用いているlen(wb1.sheetnames)は、全シート枚数。
        # なので、「最末尾シートの後ろに、新たにシートを作成する」という意味
        tgt_sh= wb2.create_sheet("読み込み不可　エラーメッセージ",len(wb2.sheetnames))
        # ↓　列幅変更　参照：https://pg-chain.com/python-excel-height-width
        tgt_sh.column_dimensions['B'].width= 20
        #n=len(alt_data)
        n=2
        #pprint.pprint('alt_data={}'.format(alt_data)) 
        for l5 in alt_data:
            for l6 in l5:
                if l6==l5[0]:
                    tgt_sh.cell(n, 2).value =l6
                elif str(l6[0])=='':
                    tgt_sh.cell(n, 2).value =''
                else:   
                    tgt_sh.cell(n, 2).value ='【'+str(l6[0])+'】'+l6[1]
                n += 1
    
    return alt_data

def koukikourei_No_Sort(ldD_obj,wb1):
    
    Session = sessionmaker(bind=engine)
    ses = Session()
    # ↓　listInsurer/後期高齢者の保険者名のリスト
    listInsurer=[]
    # ↓　InsurerDataテーブルより、…後期高齢者医療広域連合のlike検索に一致した保険者情報を、
        # ['新潟県後期高齢者医療広域連合','山形県後期高齢者医療広域連合',・・・]
        # のようにlistInsurerリストにぶち込んでいく
    # 参考：http://scm.zoomquiet.top/data/20190608075720/index.html
    
    for lIns in ses.query(InsurerData).\
        filter(InsurerData.kanji_Insurer_Name.like('%後期高齢者医療広域連合%')):
        listInsurer.append(lIns.kanji_Insurer_Name)
    ses.close()
    
    # ↓　list3/はorマ・記号・番号・患者名・シート名　の順で並ぶ　リストinリスト
    
    list3=[]
    # ちなみに、ldD_objはCulculateテーブルから引っ張り出したデータ（辞書inリスト）
    for l3 in ldD_obj:
        if '後期高齢者医療広域連合' in l3['kanji_Insurer_Name']:
            list3.append([l3['kanji_Insurer_Name'],l3['title_AcupOrMass'],l3['insuraCodeNo_Str'],l3['name'],l3['sheetName']])
    list3.sort()
    # ↓　第２引数で用いているlen(wb1.sheetnames)は、全シート枚数。
    # なので、「最末尾シートの後ろに、新たにシートを作成する」という意味
    tgt_sh= wb1.create_sheet("後期高齢者　並べ順",len(wb1.sheetnames))
    tgt_sh.cell(1, 1).value ="順番" 
    tgt_sh.cell(1, 2).value ="記号・番号"
    tgt_sh.cell(1, 3).value = "名前"
    tgt_sh.cell(1, 4).value = "シート名"
    # ↓　列幅変更　参照：https://pg-chain.com/python-excel-height-width
    tgt_sh.column_dimensions['A'].width= 5
    tgt_sh.column_dimensions['B'].width= 12
    tgt_sh.column_dimensions['C'].width= 16
    tgt_sh.column_dimensions['D'].width= 25

    x=1
    list4=['はりきゅう','マッサージ']
    # ↓　後期高齢者医療の保険者のリスト(listInsurer)が、
    # list3内の保険者と一致するものだけリスト化したものが、listInsurer2
    listInsurer2=[]
    for li3 in list3:
        for liI in listInsurer:
            if li3[0] == liI and not li3[0] in listInsurer2:
                listInsurer2.append(liI)
    # ↓　保険者が入れ替わるたびに、1行ずれて（x += 1）そこに
    # 「新潟県後期高齢者医療広域連合」などの保険者名が入る。    
    for liI2 in listInsurer2:
        x += 1
        tgt_sh.cell(x, 1).value =liI2 
        # ↓　'はりきゅう'or'マッサージ'が入れ替わるたびに、1行ずれて（x += 1）
        # そこに'はりきゅう'or'マッサージ'が入る。
        # そのたびに順番の数字（y）はリセットされる

        for li4 in list4:
            x += 1
            tgt_sh.cell(x, 1).value =li4
            y=1
            for li3 in list3:
                if li3[0] == liI2 and li3[1] == li4:
                    x += 1
                    tgt_sh.cell(x, 1).value =y
                    tgt_sh.cell(x, 2).value =li3[2]
                    tgt_sh.cell(x, 3).value =li3[3]
                    tgt_sh.cell(x, 4).value =li3[4]
                    y += 1
                
    #pprint.pprint('list3={}'.format(list3)) 
    return

def loadD_obj_furiwake_kenshikai(loadD_obj,target_sheet):
    for loadD in loadD_obj:
            #app.logger.info('insurer_No_Str={}'.format(loadD['kanji_Insurer_Name']+loadD['insurer_No_Str'][0:1]))
            if '協会' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(7, 2)#協会けんぽのはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(7, 4)#協会けんぽのはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(7, 6)#協会けんぽのマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(7, 7)#協会けんぽのマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif '共済' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(9, 2)#共済のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(9, 4)#共済のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(9, 6)#共済のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(9, 7)#共済のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif '国民健康保険組合' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(10, 2)#国保組合のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(10, 4)#国保組合のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(10, 6)#国保組合のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(10, 7)#国保組合のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif '保険組合' in loadD['kanji_Insurer_Name'] :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(8, 2)#保険組合のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(8, 4)#保険組合のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(8, 6)#保険組合のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(8, 7)#保険組合のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            #保険者番号が6桁もしくは、山形県のように5桁の場合、
            # あるいは67から始まる退職者医療の場合⇒国保へ分類される
            #なお、国保組合も、6桁～5桁だが、前elif節
            # (elif '国民健康保険組合' )のところで
            #スクリーニングしてあるので心配いらない
            elif len(loadD['insurer_No_Str']) == 6 or len(loadD['insurer_No_Str']) == 5 or loadD['insurer_No_Str'][0:2] == '67' :
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(12, 2)#国保のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(12, 4)#国保のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(12, 6)#国保のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(12, 7)#国保のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
            elif loadD['insurer_No_Str'][0:2] == '39':
                if loadD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(13, 2)#後期高齢のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(13, 4)#後期高齢のはりきゅうの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)
                elif loadD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(13, 6)#後期高齢のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(13, 7)#後期高齢のマッサージの費用額を入れるセル
                    loadDInt = int(float(loadD['amount_Str']))#費用額
                    kingaku_insert(loadDInt,target_sheet_cell2)

def KentanD_obj_furiwake_kenshikai(KentanD_obj,target_sheet):
    for kenD in KentanD_obj:
            #app.logger.info('insurer_No_Str={}'.format(kenD['kanji_Insurer_Name']+kenD['insurer_No_Str'][0:1]))
            if kenD['title_kentan'] =='県障':
                if kenD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(26, 2)#県障のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(26, 4)#県障のはりきゅうの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
                elif kenD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(26, 6)#県障のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(26, 7)#県障のマッサージの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
            elif kenD['title_kentan'] =='県老' :
                if kenD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(27, 2)#県老のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(27, 4)#県老のはりきゅうの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
                elif kenD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(27, 6)#県老のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(27, 7)#県老のマッサージの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
            elif kenD['title_kentan'] =='県親' :
                if kenD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(28, 2)#県親のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(28, 4)#県親のはりきゅうの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
                elif kenD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(28, 6)#県親のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(28, 7)#県親のマッサージの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
            elif kenD['title_kentan'] =='単子' :
                target_sheet.cell(29, 1).value='その他\n（子ども）'
                if kenD['title_AcupOrMass'] == 'はりきゅう':
                    target_sheet_cell1=target_sheet.cell(29, 2)#単子のはりきゅうの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(29, 4)#単子のはりきゅうの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
                elif kenD['title_AcupOrMass'] == 'マッサージ':
                    target_sheet_cell1=target_sheet.cell(29, 6)#単子のマッサージの件数を入れるセル
                    kensuu_insert(target_sheet_cell1)
                    target_sheet_cell2=target_sheet.cell(29, 7)#単子のマッサージの費用額を入れるセル
                    kenDInt = int(float(kenD['amount_Str']))#費用額
                    kingaku_insert(kenDInt,target_sheet_cell2)
            


