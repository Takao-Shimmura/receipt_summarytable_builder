#!python3.9.1

#　1⃣herokuにデプロイした時に、
# 　　●多ユーザー同時接続による変数のバッティング
# 　　●データベースの反応が遅いために、calculateテーブルや
# 　　　error_msgテーブルからの書き込み/読み込みエラー
#　　などが生じたために、それを解消するため
#     ●sessionの導入
#     ●dataframeデータを2重にjson化して、フロントエンド側（index2）にいったん預ける
#     ●calculate/error_msgテーブルを廃止して、変数に込める
#    などの対策を講じた。
# 　
#　2⃣構成ファイルはapp.py--index2.html--myutil2.py。
#       なお、calculate/error_msgテーブルを用いる手法は後々参考になるので、
#       apppandas.py--index.html--myutil.pyに残してある

#　3⃣メインのプログラムはapp.pyなので、デプロイ前にProcfileを確認しておく。
#       web: gunicorn app:app　の前の方のappがapp.pyのことを指す

#  4⃣ローカルで動く環境と、デプロイ時の環境で変えなきゃならないところは、
#  ### で印をつけてある（データベースのURIと最後の行のlocalhost）

from flask import Flask, render_template, request, session, \
    redirect, jsonify, current_app, g,send_file
#import psycopg2
import openpyxl # 外部ライブラリ　pip install openpyxl
import pandas as pd
#import numpy as np
import sqlite3
import json
from datetime import datetime

import os
import pathlib
import pytz
import pprint
import copy #　リストや辞書などのミュータブル（更新可能）オブジェクトをコピーする際に必要
# ↑　参照　https://note.nkmk.me/python-copy-deepcopy/

from sqlalchemy import create_engine, Column, Integer, String, \
    Text, DateTime, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy.sql.elements import Null

from myutil3 import Search_condition,InsurerData,\
    get_dic_schCond2calAttr,get_search_condition,\
        get_cellno_2list,define_soukatsu1Desti,sort_insureName_4Sokatsu1_fromloadD_obj,\
        soukatsu1Desti_List_set,get_soukatsu1Desti_insur_dic,\
        kensuu_insert,kingaku_insert,koukikourei_No_Sort,\
        error_Msg_Sheet,loadD_obj_furiwake_kenshikai,\
        KentanD_obj_furiwake_kenshikai,get_insurerData_all,\
        name_delite_space,my_round


#　↓　herokuにデプロイすると、画像ファイルが読み込めない。
# これを解消するためにflaskがheroku内でインスタンス化される時に、
# 静的なファイルのディレクトリを記述して明確化する。
# 参考⇒https://qiita.com/go_new_innov/items/222a3ed92f5ed093f462
app = Flask(__name__,static_folder='./static')
#　上記でも解消されなかった！
# ⇒原因は、画像ファイルの拡張子jpgが、
# 大文字だったため。（Xnviewで編集したため)。小文字に直したら解消した


app.secret_key = b'random string...'

#　↓　herokuのpostgreSQL接続用URI 
# ※ただし、割り当てられたURIそのままでは接続エラー
#　「postgres://・・・」から「postgresql://・・・」に変更しなければ解消されない
#参考（heroku公式リファレンス）⇒Why is SQLAlchemy 1.4.x not connecting to Heroku Postgres? - Heroku Help
engine = create_engine('postgresql://qrnkdpytaiifps:7b728dc1e568e2d1c1ab80c919e17d10c7f41f8d853c8e5989d907c978bf8d8c@ec2-34-250-16-127.eu-west-1.compute.amazonaws.com:5432/d77prcb2vt5pne')

#　↓　ローカルのSQLite接続用パス 
###engine = create_engine('sqlite:///sample.sqlite3')

# access top page.
@app.route('/',methods=['GET'])
def index():
     #'soukatsuTemp.xlsx'以外のエクセルファイルがもしサーバー上に
     #残っていたときに、HPを更新した時に予め削除してリセットしておく
    path = pathlib.Path("./")   
    for pass_obj in path.iterdir():
        if pass_obj.match("*.xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
            pass_obj.unlink()
    # session['user_access_time']は、ユーザーごとに、
    # トップページにアクセスしてきた日時を振り分けて、IDの様に用いるためのもの。
    # このsession情報は、ユーザーのPCのcookieに保存され、
    # HPが閉じられるなどして、接続が終わるまで保存され続ける。
    # 再び、トップページに戻ると、新しいアクセス日時に更新されてリセットされる。
    # 参考⇒「Python フレームワークFlaskで学ぶ Webアプリケーションのしくみとつくり方」P109        
    now = datetime.now(pytz.timezone('Asia/Tokyo'))
    session['user_access_time']=str(now.month).zfill(2)+'月' \
        +str(now.day).zfill(2) +'日'+ str(now.hour).zfill(2)+'時' \
             + str(now.minute).zfill(2) +'分'+str(now.second).zfill(2)+'秒' \
                +str(now.microsecond)+'マイクロ秒'
    d = get_insurerData_all()
    # ↑　get_insurerData_all() 保険者番号一覧をゲット
    return render_template('index3.html',\
            title = '新潟県鍼灸マッサージ師会　公認',\
            message = '保険申請書　総括票作成　ホームページ',\
            insdata = d)


# アップロード機能
@app.route('/upload', methods=['POST'])
def upload():
        df_new={}
        #df_new={}・・・pandasを用いてエクセルを読み込んで作成された
        #dataframeを、{「シート名」:「dataframe」}  という形の辞書として
        # 整理したものを入れておく変数
        parsonal_data={}
        # ↓　このif節は・・・
        # 申請の年・月・施術者名・施術所名・登録記号番号を入力する
        # ダイアログを通過した場合は、dialog_flg＝True ＞＞よってif節以下は実行されない
        # 通過していない場合（ファイルのアップデートの時）には　dialog_flg = NoneもしくはFalse
        # ＞＞よってif節以下は実行される
        flg1 = request.form.get('dialogFlg')
        #app.logger.info('flg1={}'.format(flg))
        # ↓　flg2は県単のダイアログを表示したか否かのフラグ変数
        flg2= request.form.get('kentanDialogFlg')
        if flg1 =='False':
            
            #https://blog.imind.jp/entry/2020/01/25/032249
            #を参照 拡張子チェック機能１↓ (Excelファイルの拡張子であることを確認するための仕込み段階１)
            ALLOWED_EXTENSIONS = ['.xlsx']
            # ↓ 参照元のサンプルコードでは以下の様だったが、flask.がエラーとなったため削除
            #if 'file' not in flask.request.files:
            if 'file' not in request.files:
                parsonal_data['failed_msg']='読み込めないファイル形式です　アップロード失敗'
                return jsonify(parsonal_data)

            # fileの取得（FileStorage型で取れる）
            # https://tedboy.github.io/flask/generated/generated/werkzeug.FileStorage.html
            
            # ↓ 参照元のサンプルコードでは以下の様だったが、flask.がエラーとなったため削除
            #fs = flask.request.files['file']
            fs = request.files['file']

            # 下記のような情報がFileStorageからは取れる⇒デバックコンソールに表示される仕組みにしてある
            #app.logger.info('file_name={}'.format(fs.filename))
            #app.logger.info('content_type={} content_length={}, mimetype={}, mimetype_params={}'.format(
                #fs.content_type, fs.content_length, fs.mimetype, fs.mimetype_params))
            
            #拡張子チェック機能２↓(Excelファイルの拡張子であることを確認するための仕込み段階２)
            suffix = pathlib.Path(fs.filename).suffix
            #拡張子チェック機能３↓↓(Excelファイルの拡張子であることを確認する段階)
            if not suffix in ALLOWED_EXTENSIONS:
                parsonal_data['failed_msg']="保存できないファイル形式です {}".format(suffix)
                return jsonify(parsonal_data)
            else:
                """ # ファイルを保存
                fs.save(fs.filename) """
                # ファイルを保存
                fs.save(session['user_access_time']+".xlsx")
                # ↓以下はエクセルを読み込んで、データベースに登録する段取り
            
            path = pathlib.Path("./")    #相対パス指定
            for pass_obj in path.iterdir():
                if pass_obj.match(session['user_access_time']+".xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
                    #  Pandasを用いてpd.read_excelで読み取られたエクセルの情報は、
                    #{「シート名」:「dataframe」,「シート名」:「dataframe」}  という形の辞書として取り出される。
                    # そのままdfという変数に辞書として入れておいてもいいのだが、
                    #　ユーザーID代わりのsession['user_access_time']をキーとして
                    #  {session['user_access_time']:{「シート名」:「dataframe」},...}
                    # という辞書in辞書の形で変数dfに入れ込んでおく。
                    # そうすることで、多数のユーザーが同時にアクセスしたときに、dfの中身
                    # が勝手に書き換えられたり、バッティングすることを防ぐため
                    df={}
                    df[session['user_access_time']] = pd.read_excel(pass_obj,sheet_name = None,header=None,index_col=None)
                    # ↓ アップロードされたファイルを、情報を読み取った後に削除
                    # 参考　https://www.atmarkit.co.jp/ait/articles/1910/29/news019_2.html
                    # pathlibライブラリを用いたテクニック。
                    pass_obj.unlink()

                    #　↓　df[session['user_access_time']]内にある、各シートから読み込んだ
                    # dataframeのインデックスとヘッダーを番号振りなおしして
                    # 変数df_newに入れ込んでいく。
                    # この時も、多ユーザー同時接続のバッティングを防ぐために、
                    # ユーザーID代わりのsession['user_access_time']をキーとして
                    # 格納しておく
                    df_new[session['user_access_time']]={}
                    for dfsh in df[session['user_access_time']]: 
                        dfdic=df[session['user_access_time']][dfsh]
                        dfdic.reset_index(drop=True, inplace=True)
                        shp=dfdic.shape
                        dfdic.index=range(1,shp[0]+1)
                        dfdic.columns=range(1,shp[1]+1)
                        df_new[session['user_access_time']][dfsh]=dfdic
     
        #　↓　変数condDictに、検索条件の辞書を込める
        condDict = get_search_condition()
        #app.logger.info('condDict={}'.format(condDict))
        #　↓　変数sC2cAdicに、辞書を込める
        # ’キー’は'seardhテーブル'の「属性」の文字列：
        # ’値’は’calculateテーブル’の「属性」の文字列
        sC2cAdic = get_dic_schCond2calAttr()
        #　↓　 year_month Dialogから送られてきた変数を、読み込む
        # どういうわけか、ajax通信で送られてきたものは、すべてstring型になってしまうらしい
        year_f = request.form.get('year_fixed')
        month_f = request.form.get('month_fixed')
        therapistName_f = request.form.get('therapistName_fixed')
        treatmentHosName_f = request.form.get('treatmentHosName_fixed')
        registerNo_Str_f = request.form.get('registerNo_Str_fixed')
        df_new2_f = request.form.get('df_new2')
        # ↑　df_new2_fは、一時的にフロントエンド側（index2）に送っておいた
        # dataframeの内容が、返却されてきたもの。
        # 2重にjson化されているので、それぞれjsonファイルを読み込み、
        # 最後に、session['user_access_time']をキーとした辞書にぶち込み
        # 変数df_newに格納して、後に使う
        #　json.loads()「sがついている」はjson.load()と違うことに注意！
        # 参考⇒https://note.nkmk.me/python-json-load-dump/
        # 参考⇒https://www.python.ambitious-engineer.com/archives/617
        # pd.read_json()
        # 参考⇒https://note.nkmk.me/python-pandas-to-json/
        if df_new2_f:
            df_new={}
            df_new[session['user_access_time']]={}
            df_new2=json.loads(df_new2_f)
            for k, v in df_new2.items():
                df_new[session['user_access_time']][k]=pd.read_json(v)
        #app.logger.info('df_new after={}'.format(df_new)) 
        """ wsh_id_4calc = 1 # loadD_objに乗せるデータのidをリセット
        wsh_id_4err = 1 # ErrD_objに乗せるデータのidをリセット
        wsh_id_4ken = 1 # 県単に乗せるデータのidをリセット
        wsh_id_4kenErr = 1 # 県単に乗せるエラーのデータのidをリセット """
        loadD_obj=[]
        ErrD_obj=[]        
        KentanD_obj=[]
        ErrKentanD_obj=[]
        # 先だって、フロントエンド側に送信した上記4つのデータオブジェクト
        # 達が、kentanDialogの入力を経て、バックエンド側に返ってくる。
        # その時に、json化されたものを、またリストのオブジェクトに変換
        # しなければならないので、json.loads()を用いる
        # json.loads()「sがついている」はjson.load()と違うことに注意！
        # 参考⇒https://note.nkmk.me/python-json-load-dump/
        # 参考⇒https://www.python.ambitious-engineer.com/archives/617
        if request.form.get('kentanData') != None and flg2 =='True':
            loadD_obj=json.loads(request.form.get('uploadedData'))
            ErrD_obj=json.loads(request.form.get('uploadedErrData'))
            KentanD_obj=json.loads(request.form.get('kentanData'))
            ErrKentanD_obj=json.loads(request.form.get('kentanErrData'))
        else:
            for cD in condDict:
                for dfN_Key in df_new[session['user_access_time']]:#dfN_Keyはシート名
                    df_value=df_new[session['user_access_time']][dfN_Key]#df_valueに1シート分のdataframeを入れておく。
                    # ↓　DataFrameがある大きさを越えないと、読み込まないようにしておく（はorマ　の申請用紙以外のDataFrameを読み込まない）
                    # df_value.shape[0] >= 78 は県障（新潟市内）、df_value.shape[1] >=68 は県障（新潟市外）
                    if df_value.shape[0] >= 78 and df_value.shape[1] >= 68 : 
                        if df_value.loc[get_cellno_2list(cD['acupOrMass_Cell'])[0],\
                            get_cellno_2list(cD['acupOrMass_Cell'])[1]] == cD['acupOrMass_Condition']:
                            #　↓　変数d_dicは辞書。後に一気にcalculateテーブルを更新するためのデータを入れとく
                            d_dic={}
                            d_dic['sheetName'] =dfN_Key # シート名を入れておく
                            
                            if cD['acupOrMass_Condition']=='県単医療費助成申請書' :
                                d_dic['title_kentan'] ='県障'
                                d_dic['kankatsu_kentan'] ='新潟県'
                            elif cD['acupOrMass_Condition']=='重度心身障がい者医療費助成申請書':
                                d_dic['title_kentan'] ='県障'
                                d_dic['kankatsu_kentan'] ='新潟市'
                            else:
                                d_dic['title_kentan'] ='Not県障'
                            d_dic['title_AcupOrMass'] =cD['title_AcupOrMass']
                            # ↑　はきorマを入れておく
                            # ↑　県単の場合は''空欄を入れておいて、
                            # 後にkentanダイアログにて　「県老/県障/県親/単子_はき/マ」　を入れてもらう
                            
                            # ↓「年」が入力されているセル3か所の値がnanではないときに、以下の処理を行う
                            #nanだったときには'False'
                            if not pd.isnull(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                get_cellno_2list(cD['yearTop_Cell'])[1]] ) and \
                                not pd.isnull(df_value.loc[get_cellno_2list(cD['year1st_Cell'])[0],\
                                get_cellno_2list(cD['year1st_Cell'])[1]]) and\
                                not pd.isnull(df_value.loc[get_cellno_2list(cD['yearLast_Cell'])[0],\
                                get_cellno_2list(cD['yearLast_Cell'])[1]]):

                                # ↓「年」が入力されているセル3か所の、値が一致して、なおかつ　0ではないときに
                                # 　d_dic辞書に　year_Strをキーとして、yearTop_Cellの値を込めておく
                                if int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]] )) == \
                                    int(float(df_value.loc[get_cellno_2list(cD['year1st_Cell'])[0],\
                                    get_cellno_2list(cD['year1st_Cell'])[1]]))  and\
                                    int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]]))  == \
                                    int(float(df_value.loc[get_cellno_2list(cD['yearLast_Cell'])[0],\
                                    get_cellno_2list(cD['yearLast_Cell'])[1]]))  and\
                                    int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                    get_cellno_2list(cD['yearTop_Cell'])[1]]))  != 0:
                                # ↓「年」がyear month dialogで確認した数字とあっていなければ'False'を入力        
                                    if flg1 !='False':
                                        #app.logger.info('year_f={}'.format(year_f))
                                        #app.logger.info('year={}'.format(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                        #get_cellno_2list(cD['yearTop_Cell'])[1]] ))
                                        if int(year_f) ==\
                                        int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                        get_cellno_2list(cD['yearTop_Cell'])[1]])) :
                                            d_dic['year_Str'] =\
                                            str(int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                            get_cellno_2list(cD['yearTop_Cell'])[1]] )))
                                        else:
                                            d_dic['year_Str'] ='False'
                                    else:
                                        d_dic['year_Str'] =\
                                        str(int(float(df_value.loc[get_cellno_2list(cD['yearTop_Cell'])[0],\
                                        get_cellno_2list(cD['yearTop_Cell'])[1]] )))
                                else:
                                    d_dic['year_Str'] ='False'
                            else:
                                d_dic['year_Str'] ='False'
                            # ↓「月」が入力されているセル3か所の値がnanではないときに、以下の処理を行う
                            #nanだったときには'False'
                            if not pd.isnull(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                get_cellno_2list(cD['monthTop_Cell'])[1]] ) and \
                                not pd.isnull(df_value.loc[get_cellno_2list(cD['month1st_Cell'])[0],\
                                get_cellno_2list(cD['month1st_Cell'])[1]]) and\
                                not pd.isnull(df_value.loc[get_cellno_2list(cD['monthLast_Cell'])[0],\
                                get_cellno_2list(cD['monthLast_Cell'])[1]]):    
                                # ↓「月」が入力されているセル3か所の、値が一致して、なおかつ　0ではないときに
                                # 　d_dic辞書に　month_Strをキーとして、monthTop_Cellの値を込めておく
                                if int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                    get_cellno_2list(cD['monthTop_Cell'])[1]] )) == \
                                    int(float(df_value.loc[get_cellno_2list(cD['month1st_Cell'])[0],\
                                    get_cellno_2list(cD['month1st_Cell'])[1]]))  and\
                                    int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                    get_cellno_2list(cD['monthTop_Cell'])[1]]))  == \
                                    int(float(df_value.loc[get_cellno_2list(cD['monthLast_Cell'])[0],\
                                    get_cellno_2list(cD['monthLast_Cell'])[1]]))  and\
                                    int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                    get_cellno_2list(cD['monthTop_Cell'])[1]]))  != 0:
                                # ↓「月」がyear month dialogで確認した数字とあっていなければ'False'を入力    
                                    if flg1 !='False':
                                        if int(month_f) ==\
                                        int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                        get_cellno_2list(cD['monthTop_Cell'])[1]] )):
                                            d_dic['month_Str'] =\
                                            str(int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                            get_cellno_2list(cD['monthTop_Cell'])[1]] )))
                                        else:
                                            d_dic['month_Str'] ='False'
                                    #str(int(float・・・とややこしい処理をしているのは、1⃣pandasで
                                    #読み込む際に、数値を勝手に「浮動小数点：float」で読み込んで
                                    # dataframe化されるがある（例「2」のはずが「2.0」と読み込む）
                                    # そのために、2⃣そのデータがstrによって文字列化されてしまうと、
                                    # (例「'2.0'」)3⃣それをさらに、intで整数化しようとするとエラーが出る
                                    #参考⇒https://qiita.com/ringCurrent/items/1df058bb203374a4b294
                                    #これらを回避するために、float関数を用いる


                                    else:
                                        d_dic['month_Str'] =\
                                        str(int(float(df_value.loc[get_cellno_2list(cD['monthTop_Cell'])[0],\
                                        get_cellno_2list(cD['monthTop_Cell'])[1]] )))
                                else:
                                    d_dic['month_Str'] ='False'
                            else:
                                d_dic['month_Str'] ='False'

                            for sC,cA in sC2cAdic.items():
                            # {sC:cA}の辞書は、search conditionとcalculate attributeの略。
                            #  get_dic_schCond2calAttr()関数で設定してある。
                            # month,year,name_nospace,sheetname,title_AcupOrMass以外の項目を、
                            # 検索セルと項目のセットで辞書化したもの   
                                try:
                                    # これから判定しなければならないdataframeの各セルのデータを以下の
                                    # cellV1に予め込めておく
                                    cellV1=df_value.loc[get_cellno_2list(cD[sC])[0],\
                                        get_cellno_2list(cD[sC])[1]]
                                    # pandasで取得したdataFrameのなかで、欠損値である'nan'は扱いが難しく、
                                    # if文で判定するためには、 if cA=='nan' や　if cA==str('nan')では
                                    # 判定してくれない（しかも構文エラーにならないので、ややこしい）
                                    #　判定するためにはif pd.isnull(?):とする
                                    # （欠損値nanならば’true’／pdは　import pandas as pd　より）
                                    # 参照⇒https://kagglenote.com/misc/pandas_nan_judge/
                                    if not pd.isnull(cellV1) :

                                        
                                        # ↓　もしも、更新先のテーブルの「属性」に'insurer_No_Str'(保険者番号)という文字列
                                        # が含まれていたら、'insurerNoLast_Cell'と'insurerNo_CellStep'を駆使して
                                        # 保険者番号を抽出し、'insurer_No_Str'をキーとして
                                        # 文字列として入れておく
                                        if 'insurer_No_Str' in cA:
                                            number = ''
                                            for n in range(0,8,1):
                                                
                                                #  ↓　DataFrameの場合、値が入っていない場合は’nan’
                                                #　判定はpd.isnull()
                                                if pd.isnull(df_value.loc[get_cellno_2list(cD[sC])[0],\
                                                get_cellno_2list(cD[sC])[1]+n*cD['insurerNo_CellStep']]):
                                                    jj = ''
                                                #　↑　これによって、法別番号（保険者番号の上2桁）が「なし」
                                                # の場合もOK
                                                # ↓保険者番号の1マスに、整数だけでなく、小数を含んだ数字がはいっているかもしれないので、
                                                # str(int(float(によって、無理やり整数化と文字列化をする。
                                                else:
                                                    try:
                                                        jj = str(int(float(df_value.loc[get_cellno_2list(cD[sC])[0],\
                                                        get_cellno_2list(cD[sC])[1]+n*cD['insurerNo_CellStep']] )))
                                                    #　↓　str(int(floatでエラーが出る場合は、「数値に変換できない文字列」が入っている場合
                                                    # そういう場合は、numberに'False'を入れて、breakでfor文を
                                                    # とっとと抜け出す
                                                    except:
                                                        number= 'False'
                                                        break
                                                number = jj + number
                                            d_dic[cA] = number
                                        # ↓　もしも、更新先のテーブルの「属性」が'therapistName'(施術者名)だったら
                                        # なおかつyear month dialogで確認した施術者名と違っていたら'False'が入る
                                        elif 'therapistName' in cA and flg1 !='False' and\
                                            therapistName_f != cellV1 :
                                                d_dic[cA] = 'False'
                                            
                                        # ↓　もしも、更新先のテーブルの「属性」が'treatmentHosName'(施術所名)だったら
                                        # なおかつyear month dialogで確認した施術所名と違っていたら'False'が入る
                                        elif 'treatmentHosName' in cA and flg1 !='False' and\
                                            treatmentHosName_f != cellV1 :
                                                d_dic[cA] = 'False'
                                        # ↓　もしも、更新先のテーブルの「属性」が'registerNo_Str'(登録記号番号)だったら
                                        # なおかつyear month dialogで確認した登録記号番号と違っていたら'False'が入る
                                        elif 'registerNo_Str' in cA and flg1 !='False' and\
                                            registerNo_Str_f != cellV1 :
                                                d_dic[cA] = 'False'
                                        
                                        # ↓　もしも、更新先のテーブルの「属性」が'amount_Str'(合計額)もしくは
                                        # 'copayment_Str'(一部負担金額)もしくは'billingAmount_Str'(請求額)であり、
                                        # なおかつyear_month Dialogを開き終わった後だったら・・・
                                        # cellV1の値を文字列化したものをd_dic[cA] に入れてみようとする（try文）
                                        # エラーが出る（cellV1が文字、もしくはnanだったら）exceptに飛んで、'False'が入る。
                                        # うまくいったとしても、cellV1が文字列'0'だったら'False'が入る。
                                        # for文から抜け出す　参考⇒https://note.nkmk.me/python-break-nested-loops/
                                        # カッコと論理演算子　参考⇒https://dot-blog.jp/news/python-boolean-operations-bool/
                                        elif 'amount_Str' in cA or 'copayment_Str' in cA \
                                            or 'billingAmount_Str' in cA :
                                            try:
                                                d_dic[cA] = str(int(my_round(float(cellV1))))
                                                if str(int(float(cellV1)))=='0':
                                                    d_dic[cA] = 'False'  
                                            except:
                                                d_dic[cA] = 'False'  
                                        # 患者氏名（'name_nospace' ）ならば、
                                        # セルの値からスペースを削除して 入力する
                                        elif 'name_nospace' in cA:
                                            d_dic[cA] =name_delite_space(cellV1)

                                        # 上記以外ならば、素直にセルの値が入る。
                                        else:
                                            d_dic[cA] = cellV1
                                    # 'nan'　つまり参照したセルが空白ならば'False'が入る。
                                    else:
                                        d_dic[cA] = 'False'
                                except:
                                    if cD[sC]=='pass':
                                        d_dic[cA] = 'Thru'
                            # app.logger.info('d_dic[insurer_No_Str][0:4]={}'.format(d_dic['insurer_No_Str'][0:4]))
                            define_soukatsu1Desti(d_dic)
                            
                            # ↓ valFalに一つでも'False'文字列が入っていれば、
                            # 'error_msgテーブル'に更新され、
                            # 'False'文字列が入っていなければ、'calculateテーブル'に更新される
                            # ↓　for文のbreakやelseの使い方は 
                            # 右を参照　https://python.civic-apps.com/else-loop/
                            for valFal in d_dic.values():
                                #　↓　県単のデータであることを示すものが、d_dicの値の中にあった場合
                                if valFal=='Thru' or valFal=='県障' :
                                    # ↓　もう一回、そのd_dicの値をfor文で洗いざらい調べて、
                                    # 'False'があった場合には、ErrKentanD_obj　県単のエラーデータに収める
                                    for valFal2 in d_dic.values():
                                        #　↓　flg2 !='False'⇒県単のダイアログを表示した後の場合 
                                        if valFal2=='False'and flg2 !='False':
                                            """ d_dic['id'] =wsh_id_4kenErr
                                            wsh_id_4kenErr += 1 """
                                            ErrKentanD_obj.append(d_dic)
                                            break
                                    #　↓それ以外は、valFal2のfor文を抜け出して、KentanD_obj　県単のデータに収める
                                    else:
                                        """ d_dic['id'] =wsh_id_4ken
                                        wsh_id_4ken += 1 """
                                        KentanD_obj.append(d_dic)
                                    break
                                #　↓　県単以外で'False'などが見受けられたら、ErrD_objに収める
                                elif valFal=='False'or pd.isnull(valFal) or valFal=='NotFound' or valFal=='0' or valFal=='00000000':
                                    """ d_dic['id'] =wsh_id_4err
                                    wsh_id_4err += 1 """
                                    
                                    ErrD_obj.append(d_dic)

                                    break
                            #　↓　県単以外で'False'などがまったくなかったら、 loadD_objに収める。
                            # もしくは、year month dialogに飛ぶ
                            else:
                                """ d_dic['id'] =wsh_id_4calc """
                                # year_month Dialogにて、年・月・施術者名などを確認するため、
                                # いずれの項目にもFalseがない一番目のレコードで、なおかつ
                                # year_month Dialogがまだ開かれていない(flg1 =='False')場合、
                                # 施術管理者名や登録記号番号、施術署名をparsonal_dataにぶっこんで
                                # jsonifyしてreturnで返す
                                """ if wsh_id_4calc == 1 and flg1 =='False': """
                                if len(loadD_obj) == 0 and flg1 =='False' and d_dic['title_kentan'] !='県障':
                                    parsonal_data['therapistName']=d_dic['therapistName']
                                    parsonal_data['treatmentHosName']=d_dic['treatmentHosName']
                                    parsonal_data['registerNo_Str']=d_dic['registerNo_Str']
                                    parsonal_data['year_Int'] =int(d_dic['year_Str'])
                                    parsonal_data['month_Int'] =int(d_dic['month_Str'])
                                    
                                    # ↓　year_month Dialogに遷移する前に、df_new[session['user_access_time']]
                                    # に格納したdataframeが消えてしまわないように、一時的にフロントエンド側
                                    # （index2）に送って保存しておいてもらう。
                                    # parsonal_data内に辞書として格納されたdf_new2は、jsonify(parsonal_data)
                                    # によって、一回json化されるものの、それだけではエラーが出てしまう。
                                    # なぜなら、dataframe部分は単純にjson化できないから。
                                    # つまり、dataframe部分を先に一度json化して、それをさらにもう一度全体を
                                    # jsonify(parsonal_data)で2重にjson化しなければ、フロントエンド側には送れない。 
                                    
                                    # pd.to_json()の使い方
                                    # 参考⇒https://note.nkmk.me/python-pandas-to-json/
                                    df_new2={}
                                    for k, v in df_new[session['user_access_time']].items():
                                        df_new2[k]=v.to_json()
                                    parsonal_data['df_new2']=df_new2
                                    # year_month Dialogへと遷移する
                                    return jsonify(parsonal_data)
                            # ↓　いずれの項目にもFalseがなく、year_month Dialogが
                            # すでに開かれている場合(flg1 =='False')場合、loadD_obj
                            # に加えられていく
                                """ wsh_id_4calc += 1 """
                                loadD_obj.append(d_dic)

            #--------------------------------------------------------------------
            #ここまでで、 df_new[session['user_access_time']]から
            # loadD_obj、ErrD_obj、KentanD_obj、ErrKentanD_objへの書き込みが
            # すべて終わっている状態
            #--------------------------------------------------------------------

            # ↓　kentanDialogが一度も開かれていない状態で、なおかつ
            # KentanD_objに何か値が入っている・・・つまり県単の書類が存在する場合、
            # parsonal_dataにloadD_obj~ErrKentanD_objを詰め込んで、json化して
            # index3にお返しする⇒kentanDialogにて「県老/県障/県親/単子_はき/マ」を
            # 入力してもらい、/uploadに戻ってきてもらう。
            #pprint.pprint('ErrKentanD_obj{}'.format(ErrKentanD_obj))    
            if flg2 ==None and KentanD_obj != [] :
            #if flg2 =='False' and KentanD_obj != []:
                parsonal_data['uploadedData']=loadD_obj
                parsonal_data['uploadedErrData']=ErrD_obj
                parsonal_data['kentanData']=KentanD_obj
                parsonal_data['kentanErrData']=ErrKentanD_obj
                return jsonify(parsonal_data)



        parsonal_data['process_msg']='総括票　作成中・・・' 
        #--------------------------------------------------------------------
        # 新潟県師会の総括表に県障の金額を入れ込む際に、
        # 同じ患者の療養費支給申請書から「一部負担金」の金額を持ってきて、
        # 'amount_Str'の値に入れ込む　というセクション
        #--------------------------------------------------------------------
        # ↓　いったん、KentanD_objやErrKentanD_objを別のオブジェクトとして
        # 変数KentanD_obj_kariやErrKentanD_obj_kariにコピーする。
        # 単純に　KentanD_obj_kari=KentanD_obj　等としてしまいがちだが、
        # リストや辞書などのミュータブル（更新可能）オブジェクトが
        # 代入された変数を、さらに別の変数に代入した場合、いずれかの変数を
        # 更新（要素の変更や追加・削除など）すると他方も更新される・・・という
        # 問題を回避するため、深いコピーcopy.deepcopyを使って、多次元リストを
        # コピーする。そのためには　import copy　も必要なことをお忘れなく。
        # 参照⇒https://note.nkmk.me/python-copy-deepcopy/

        KentanD_obj_kari=copy.deepcopy(KentanD_obj)
        ErrKentanD_obj_kari=copy.deepcopy(ErrKentanD_obj)

        for KentanD in KentanD_obj:
            #app.logger.info('KentanD_obj={}'.format(KentanD_obj)) 
            #app.logger.info('name_nospace_KentanD={}'.format(KentanD['name_nospace'])) 
            kentanFlg=True
            for loadD in loadD_obj:
                #　↓　loadD['name_nospace']に一致するKentanDがあった場合、
                # 　KentanD['amount_Str']にloadD['copayment_Str'] をぶち込む
                if loadD['name_nospace']==KentanD['name_nospace']:
                    kentanFlg=False
                    #KentanD['amount_Str']=loadD['copayment_Str']
                    # ↓　このコピーが無いと、↑が新潟県師会の総括票にに反映されないままになる
                    for kari in KentanD_obj_kari:
                        if kari['name_nospace']==KentanD['name_nospace']:
                            kari['amount_Str']=loadD['copayment_Str']
            #　↓　loadD['name_nospace']に一致するKentanDが無かった場合
            # つまり、県単の申請書は存在したが、それと同名患者の療養費の申請書が無かった場合
            if kentanFlg:
                # ↓　先にKentanDをKentanD_obj_kari.から削除する
                # （中身を書き換えた後だと、removeで検索できずエラーが生じる）
                KentanD_obj_kari.remove(KentanD)
                    #　↓　合計額に'NotFound'を差し替えて、後に
                    # myutil3.py のerror_Msg_Sheet()にて処理してもらう
                KentanD['amount_Str']='NotFound'
                #　↓　KentanDをErrKentanD_obj_kariに加える
                #app.logger.info('name_nospace_KentanD_beforeCHANGE={}'.format(KentanD_obj)) 
                ErrKentanD_obj_kari.append(KentanD)
                #app.logger.info('name_nospace_KentanD_midCHANGE={}'.format(KentanD_obj)) 
                
                #app.logger.info('KentanD_obj_after1={}'.format(KentanD_obj))
                #app.logger.info('KentanD_obj_after2={}'.format(KentanD_obj))
        KentanD_obj=copy.deepcopy(KentanD_obj_kari)
        ErrKentanD_obj=copy.deepcopy(ErrKentanD_obj_kari)      

        #--------------------------------------------------------------------
                
                
        #　↓　ブックの複製　参照⇒https://neko-py.com/python-excel-write-book
        wb = openpyxl.load_workbook(filename='soukatsuTemp.xlsx')
        #　↓　日付や時間の取得　参照⇒https://www.sejuku.net/blog/23606
        # しかし、上記のとおりに、date = datetime.datetime.now()　と書くとエラー
        date = datetime.now()
        # 2桁表示のゼロパディングは　参照⇒https://note.nkmk.me/python-zero-padding/
        # loadD_obj)のから総括票ⅠⅡを作るべく、保険者情報を順番に並べなおしたのが、sortInsList
        sortInsList=sort_insureName_4Sokatsu1_fromloadD_obj(loadD_obj)
        template_sheet = wb['総括票（Ⅰ）(ひな形　禁削除)']
        # ↓　総括表１の送付先（soukatsu1Desti）だけを、重複なくリスト化したものがsoukatsu1Desti_List
        soukatsu1Desti_List = soukatsu1Desti_List_set(sortInsList)


        for desti in soukatsu1Desti_List:
            dicDesti_insur= get_soukatsu1Desti_insur_dic(sortInsList,soukatsu1Desti_List) 
            #app.logger.info('desti={}'.format(desti))  
            #app.logger.info('count1={}'.format(counter1))                     
            #app.logger.info('count1int={}'.format((int(counter1 / 7 - 0.1)+1)))
            # ↓同じ総括表１の行き先（soukatsu1Desti）に、どれだけの保険者の数がはいるか？
            # を、変数yyに込める
            if '△' not in desti:# 総括票Ⅰを作りたくない保険者
                yy=int(len(dicDesti_insur[desti]) / 7 - 0.1)+1
                for x in range(yy):     
                        target_sheet = wb.copy_worksheet(template_sheet)
                        #app.logger.info('template_sheet={}'.format(template_sheet.sheet_properties.tabColor))
                        # ↓　複製したシートのタブの色を、色なしにする
                        target_sheet.sheet_properties.tabColor =None
                        if x >= 1:
                            target_sheet.title = '総括票（Ⅰ）('+ desti +'）'+str(x+1)+'枚目'
                        else: 
                            target_sheet.title = '総括票（Ⅰ）('+ desti +'）'
                        # ↓ 複製したシート(総括票（Ⅰ）)に、それぞれの数値を入れる
                        target_sheet.cell(14, 26).value = registerNo_Str_f #登録記号番号
                        target_sheet.cell(17, 26).value = therapistName_f #施術管理者
                        target_sheet.cell(20, 26).value = treatmentHosName_f #施術所名
                        target_sheet.cell(6, 10).value = year_f #年
                        target_sheet.cell(6, 21).value = month_f #月
                        # ↓ 複製したシートが複数にわたり、x枚目がその最後の時は、
                        # 保険者名が7段目までいかずに、途中(y段目)で終わるような仕掛け
                        if x == yy-1:
                            y = len(dicDesti_insur[desti]) - 7*(x)
                        # ↓ 複製したシートが複数にわたり、x枚目がその途中の時は、
                        # 7段目までフルに入力する
                        elif x < (yy-1):
                            y = 7
                        #app.logger.info('y={}'.format(y))
                        #app.logger.info('target_sheet.title={}'.format(target_sheet.title))
                        
                        # ↓ 複製したシート（総括表１）に、保険者名＋改行＋（はりきゅうorマッサージ）
                        # を、上から順にｖ番目まで入れていく
                        # （数列v+7*xを用いて、コピーがx枚目のときは、1枚目の続きの保険者が入るようにしてある）
                        
                        # ↓ 金額を変数に入力していく際に、my_round()というオリジナル関数（myutil3内で定義）
                        # して、文字列⇒float浮動小数点に変換された数値を、小数点以下を四捨五入して、
                        # さらにint関数によって整数化している。（20220830修正）

                        v=0
                        for v in range(y):
                            listv = dicDesti_insur[desti]
                            target_sheet.cell(32 + 6 * v, 2).value = listv[v+7*x][0]+'\n'+'('+listv[v+7*x][1]+')'
                            
                            for loadD in loadD_obj:
                                if loadD['kanji_Insurer_Name'] == listv[v+7*x][0] \
                                    and loadD['title_AcupOrMass'] == listv[v+7*x][1]:
                                    if loadD['relationship'] == '本人':
                                        target_sheet_cell1=target_sheet.cell(34 + 6 * v, 15)#本人の件数を入れるセル
                                        kensuu_insert(target_sheet_cell1)
                                        target_sheet_cell2=target_sheet.cell(34 + 6 * v, 21)#本人の費用額を入れるセル
                                        loadDInt = int(my_round(float(loadD['amount_Str'])))#本人の費用額
                                        kingaku_insert(loadDInt,target_sheet_cell2)
                                    else:
                                        target_sheet_cell1=target_sheet.cell(34 + 6 * v, 33)#家族の件数を入れるセル
                                        kensuu_insert(target_sheet_cell1)
                                        target_sheet_cell2=target_sheet.cell(34 + 6 * v, 39)#家族の費用額を入れるセル
                                        loadDInt = int(my_round(float(loadD['amount_Str'])))#家族の費用額
                                        kingaku_insert(loadDInt,target_sheet_cell2)
        template_sheet = wb['総括票（Ⅱ）(ひな形　禁削除)']            
        for insurL in sortInsList:
            target_sheet = wb.copy_worksheet(template_sheet)
            target_sheet.sheet_properties.tabColor =None
            target_sheet.title = '総括票（Ⅱ）' + insurL[3] + '(' + insurL[2][0] +  ')'
            target_sheet.cell(18, 26).value = registerNo_Str_f #登録記号番号
            target_sheet.cell(21, 26).value = therapistName_f #施術管理者
            target_sheet.cell(24, 26).value = treatmentHosName_f #施術所名
            target_sheet.cell(6, 10).value = year_f #年
            target_sheet.cell(6, 21).value = month_f #月
            target_sheet.cell(14, 15).value = insurL[3]+'\n'+'('+insurL[2] +')' #保険者名＋はorマ
            for loadD in loadD_obj:
                if loadD['kanji_Insurer_Name'] == insurL[3] \
                    and loadD['title_AcupOrMass'] == insurL[2]:
                    if loadD['relationship'] == '本人':
                        target_sheet_cell1=target_sheet.cell(35, 13)#本人の件数を入れるセル
                        kensuu_insert(target_sheet_cell1)
                        target_sheet_cell2=target_sheet.cell(35, 24)#本人の費用額を入れるセル
                        loadDInt = int(my_round(float(loadD['amount_Str'])))#本人の費用額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(35, 39)#本人の一部負担金額を入れるセル
                        loadDInt = int(my_round(float(loadD['copayment_Str'])))#本人の一部負担金額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(35, 54)#本人の請求額を入れるセル
                        loadDInt = int(my_round(float(loadD['billingAmount_Str'])))#本人の請求額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                    else:
                        target_sheet_cell1=target_sheet.cell(41, 13)#家族の件数を入れるセル
                        kensuu_insert(target_sheet_cell1)
                        target_sheet_cell2=target_sheet.cell(41, 24)#家族の費用額を入れるセル
                        loadDInt = int(my_round(float(loadD['amount_Str'])))#家族の費用額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(41, 39)#家族の一部負担金額を入れるセル
                        loadDInt = int(my_round(float(loadD['copayment_Str'])))#家族の一部負担金額
                        kingaku_insert(loadDInt,target_sheet_cell2)
                        target_sheet_cell2=target_sheet.cell(41, 54)#家族の請求額を入れるセル
                        loadDInt = int(my_round(float(loadD['billingAmount_Str'])))#家族の請求額
                        kingaku_insert(loadDInt,target_sheet_cell2)
        
        template_sheet = wb['総括表　新潟県師会用（禁削除）']            
        target_sheet = wb.copy_worksheet(template_sheet)
        target_sheet.sheet_properties.tabColor =None
        # ↓　.xlsxファイルが作成された年月日と時刻を、日本時間で取得
        # （サーバーのおかれている国によって、時刻が変動しないように）
        # 参考⇒https://qiita.com/keisuke0508/items/df2594770d63bf124ccd
        now = datetime.now(pytz.timezone('Asia/Tokyo'))

        target_sheet.title = '総括表　新潟県師会用('+str(now.month).zfill(2)+'月' +str(now.day).zfill(2) +'日'+ str(now.hour).zfill(2)+'時' + str(now.minute).zfill(2) +'分 作成'+')' 
        target_sheet.cell(2, 2).value = year_f #年
        target_sheet.cell(2, 4).value = month_f #月
        target_sheet.cell(3, 5).value = therapistName_f #施術管理者
        
        loadD_obj_furiwake_kenshikai(loadD_obj,target_sheet)
        KentanD_obj_furiwake_kenshikai(KentanD_obj,target_sheet)
        
        wb.remove(wb['総括表　新潟県師会用（禁削除）'])
        wb.remove(wb['総括票（Ⅰ）(ひな形　禁削除)'])
        wb.remove(wb['総括票（Ⅱ）(ひな形　禁削除)'])
        koukikourei_No_Sort(loadD_obj,wb)
        
        parsonal_data['alert_data'] = error_Msg_Sheet(ErrD_obj,ErrKentanD_obj,wb)
        #app.logger.info('parsonal_data[alert_data] ={}'.format(parsonal_data['alert_data'] ))  
        #app.logger.info('dicDesti_insur={}'.format(dicDesti_insur))

        
        # ↓　作成された.xlsxファイルに、作成年月日でファイル名を命名する
        # ↓　時刻の2桁表示（ゼロ埋め）は.zfill()で行う
        # 参考⇒https://note.nkmk.me/python-zero-padding/
        try:
            dLFileName='総括票 令和'+year_f+'年'+month_f+'月分　'+str(now.month).zfill(2)+'月' +str(now.day).zfill(2) +'日'+ str(now.hour).zfill(2)+'時' + str(now.minute).zfill(2) +'分'+str(now.second).zfill(2) +'秒 作成'+ '.xlsx'
        except:
            dLFileName='すべてのシートが読み込み不可'+ '.xlsx'
        
        wb.save(dLFileName)
        wb.close() 
        parsonal_data['dLFile']=dLFileName

        return  jsonify(parsonal_data)
        # '総括票 令和'+year_f+'年'+month_f+'月分　'+str(date.month).zfill(2)+'月' +str(date.day).zfill(2) +'日'+ str(date.hour).zfill(2)+'時' + str(date.minute).zfill(2) +'分'+str(date.second).zfill(2) +'秒 作成'+ '.xlsx')

@app.route('/download', methods=['POST'])
def download(): 
    #　↓　作成されたファイル名をindexのload_file()から送られてきた
    # formDataから読み取り、ファイルをsend_fileによってindexに送って
    # ダウンロードさせる
    # 参照⇒https://qiita.com/5zm/items/760000cf63b176be544c#2-%E6%96%B9%E6%B3%95%EF%BC%91send_file%E3%82%92%E5%88%A9%E7%94%A8%E3%81%99%E3%82%8B
    fName= request.form.get('filename')  
    #app.logger.info('fName={}'.format(fName)) 
    download_file_name=fName
    download_file = fName
    XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    sendingFile=send_file(download_file, as_attachment=True,\
              attachment_filename=download_file_name,\
              mimetype=XLSX_MIMETYPE)

    return sendingFile

@app.route('/dLfileDel', methods=['GET'])
def dlf_delete(): 
    path = pathlib.Path("./") 
    for pass_obj in path.iterdir():
            if pass_obj.match("*.xlsx") and pass_obj.name != 'soukatsuTemp.xlsx':
                # ↓ 作成された総括票ファイルを、ダウンロード後に削除
                # 参考　https://www.atmarkit.co.jp/ait/articles/1910/29/news019_2.html
                # pathlibライブラリを用いたテクニック。
                pass_obj.unlink()
    return '総括票を作成しました。ダウンロードして「名前を付けて保存」してください。'
if __name__=='__main__':
    app.debug = True

    # ↓　サーバーにデプロイして公開するためのもの
    app.run(host='0.0.0.0')
    # ↓　ローカルで用いるためのもの
    ###app.run(host='localhost')